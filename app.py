import csv
import io
import json
import logging
import os
import re
import secrets
import sqlite3
import textwrap
import time
import zipfile
import zlib
from collections import defaultdict
from datetime import date as date_type
from datetime import datetime, time as time_type, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps
from pathlib import Path

from flask import (
    Flask,
    Response,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
REPORT_DIR = BASE_DIR / "generated_reports"
DB_PATH = INSTANCE_DIR / "toolkit_reports.sqlite3"
LOGO_PATH = BASE_DIR / "static" / "toolkit-logo.png"

INSTANCE_DIR.mkdir(exist_ok=True)
REPORT_DIR.mkdir(exist_ok=True)
(BASE_DIR / "logs").mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(BASE_DIR / "logs" / "app.log", delay=True),
    ],
)
log = logging.getLogger(__name__)

DATABASE_URL = os.environ.get("DATABASE_URL", "")

PG_PARAM = "%s"
SQLITE_PARAM = "?"

def _make_sqlite_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _make_pg_conn():
    import psycopg2
    import psycopg2.extras
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False
    return conn


def _is_pg():
    return DATABASE_URL.startswith("postgresql")


def _param():
    return PG_PARAM if _is_pg() else SQLITE_PARAM


def _adapt(sql):
    return sql.replace("?", _param())


def _row_to_dict(cursor, row):
    if row is None:
        return None
    columns = [desc[0] for desc in cursor.description]
    return dict(zip(columns, row))


class DbCursor:
    def __init__(self, cursor, is_pg):
        self._cursor = cursor
        self._is_pg = is_pg

    @property
    def lastrowid(self):
        if self._is_pg:
            self._cursor.execute("SELECT LASTVAL()")
            return self._cursor.fetchone()[0]
        return self._cursor.lastrowid

    @property
    def description(self):
        return self._cursor.description

    def fetchone(self):
        row = self._cursor.fetchone()
        if row is None:
            return None
        if self._is_pg:
            return _row_to_dict(self._cursor, row)
        return row

    def fetchall(self):
        rows = self._cursor.fetchall()
        if self._is_pg:
            return [_row_to_dict(self._cursor, r) for r in rows]
        return rows

    def close(self):
        self._cursor.close()


class DbConnection:
    def __init__(self, conn, is_pg):
        self._conn = conn
        self._is_pg = is_pg

    def execute(self, sql, params=None):
        cur = self._conn.cursor()
        cur.execute(_adapt(sql), params or [])
        return DbCursor(cur, self._is_pg)

    def executescript(self, sql):
        if self._is_pg:
            cur = self._conn.cursor()
            for statement in sql.split(";"):
                s = statement.strip()
                if s:
                    cur.execute(s)
            cur.close()
        else:
            self._conn.executescript(sql)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()


app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "local-dev-change-me")
app.config["MAX_CONTENT_LENGTH"] = 512 * 1024
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("ENV") == "production"
SESSION_IDLE_TIMEOUT = timedelta(hours=8)
app.config["PERMANENT_SESSION_LIFETIME"] = SESSION_IDLE_TIMEOUT
app.config["REMEMBER_COOKIE_DURATION"] = timedelta(days=14)

DEFAULT_PRODUCTION_HOSTS = "reports.toolkitafrica.ac.ke"
ALLOWED_HOSTS = {
    host.strip().lower()
    for host in os.environ.get("ALLOWED_HOSTS", DEFAULT_PRODUCTION_HOSTS).split(",")
    if host.strip()
}

RATE_LIMIT_SECONDS = 30
DASHBOARD_LIST_CAP = 50  # most-recent rows shown on the dashboard; stats are not limited
LOGIN_RATE_LIMIT_WINDOW = 300
LOGIN_RATE_LIMIT_MAX = 10
ACCOUNT_LOCKOUT_THRESHOLD = 5
ACCOUNT_LOCKOUT_DURATION = timedelta(minutes=15)
INACTIVITY_LOCK_DAYS = 14

_rate_limit_store = defaultdict(list)
_login_attempts_cache = defaultdict(list)

PDF_COLORS = {
    "forest": (0.105, 0.302, 0.180),  # #1B4D2E
    "green": (0.180, 0.490, 0.275),  # #2E7D46
    "gold": (0.961, 0.769, 0.000),  # #F5C400
    "pale_green": (0.925, 0.965, 0.929),
    "line": (0.650, 0.760, 0.680),
    "ink": (0.080, 0.120, 0.095),
    "muted": (0.330, 0.410, 0.360),
    "white": (1, 1, 1),
}


def generate_csrf_token():
    if "_csrf_token" not in session:
        session["_csrf_token"] = secrets.token_hex(32)
    return session["_csrf_token"]


def validate_production_config():
    if os.environ.get("ENV") == "production" and app.config["SECRET_KEY"] == "local-dev-change-me":
        raise RuntimeError("SECRET_KEY must be set before running in production.")


def csrf_protect():
    token = session.get("_csrf_token")
    given = request.form.get("_csrf_token") or request.headers.get("X-CSRF-Token", "")
    if not token or not given or not secrets.compare_digest(token, given):
        abort(400)


def rate_limit(key_prefix, max_attempts, window_seconds, redirect_to="login"):
    def decorator(fn):
        @wraps(fn)
        def wrapped(*args, **kwargs):
            ip = request.remote_addr or "unknown"
            now = time.time()
            store_key = f"{key_prefix}:{ip}"
            timestamps = _rate_limit_store[store_key]
            cutoff = now - window_seconds
            while timestamps and timestamps[0] < cutoff:
                timestamps.pop(0)
            if len(timestamps) >= max_attempts:
                retry_after = int(timestamps[0] + window_seconds - now)
                log.warning("Rate limit exceeded for %s from %s", key_prefix, ip)
                flash(f"Too many attempts. Try again in {retry_after} seconds.", "danger")
                return redirect(url_for(redirect_to))
            timestamps.append(now)
            return fn(*args, **kwargs)
        return wrapped
    return decorator


def check_account_locked(email):
    now = time.time()
    attempts = _login_attempts_cache.get(email, [])
    cutoff = now - ACCOUNT_LOCKOUT_DURATION.total_seconds()
    while attempts and attempts[0] < cutoff:
        attempts.pop(0)
    if len(attempts) >= ACCOUNT_LOCKOUT_THRESHOLD:
        return True
    return False


def record_failed_attempt(email):
    _login_attempts_cache.setdefault(email, []).append(time.time())


def clear_login_attempts(email):
    _login_attempts_cache.pop(email, None)


def security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "0"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    if os.environ.get("ENV") == "production":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "font-src 'self' https://fonts.gstatic.com; "
            "img-src 'self' data:; "
            "base-uri 'self'; "
            "form-action 'self'"
        )
    return response


@app.before_request
def enforce_allowed_host():
    if os.environ.get("ENV") != "production" or not ALLOWED_HOSTS:
        return None

    host = request.host.split(":", 1)[0].lower()
    if host not in ALLOWED_HOSTS:
        log.warning("Blocked request for unexpected host %s", request.host)
        return Response("Misdirected request", status=421, mimetype="text/plain")


@app.before_request
def enforce_https():
    if os.environ.get("ENV") == "production" and not app.debug:
        scheme = request.headers.get("X-Forwarded-Proto", "")
        if not scheme:
            scheme = request.headers.get("X-Forwarded-Scheme", "")
        if not scheme:
            scheme = "https" if request.headers.get("HTTPS", "").lower() in ("on", "1") else ""
        if not scheme:
            scheme = request.scheme
        if scheme == "http":
            return redirect(request.url.replace("http://", "https://", 1), 301)


@app.after_request
def apply_security_headers(response):
    return security_headers(response)


@app.before_request
def check_session_expiry():
    if "user_id" in session:
        last = session.get("_last_active")
        if last:
            try:
                last_dt = datetime.strptime(last, "%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                last_dt = None
            if last_dt and (datetime.now() - last_dt) > SESSION_IDLE_TIMEOUT:
                session.clear()
                flash("Session expired due to inactivity.", "info")
                return redirect(url_for("login"))
        session["_last_active"] = now()


DEPARTMENTS = [
    "Marketing",
    "Admissions",
    "Digital Recruitment",
    "Finance",
    "Training",
    "Student Affairs",
    "Administration",
    "ICT",
    "Management",
    "Other",
]


BRANCHES = [
    "Toolkit Skills & Innovation Hub - Kikuyu",
    "Toolkit Africa Main Office",
    "Remote / Field Work",
    "Other",
]

METRICS_CONFIG = {
    "Marketing": [
        {"key": "tiktok_videos", "label": "TikTok videos", "type": "number"},
        {"key": "facebook_posts", "label": "Facebook posts", "type": "number"},
        {"key": "instagram_posts", "label": "Instagram reels/stories", "type": "number"},
        {"key": "whatsapp_followups", "label": "WhatsApp follow-ups", "type": "number"},
        {"key": "calls_made", "label": "Calls made", "type": "number"},
        {"key": "leads_generated", "label": "Leads generated", "type": "number"},
        {"key": "testimonials_collected", "label": "Testimonials collected", "type": "number"},
        {"key": "outreach_updates", "label": "Outreach updates", "type": "number"},
        {"key": "crm_updated", "label": "CRM updated?", "type": "select", "options": ["Yes", "No", "Not applicable"]},
    ],
    "Admissions": [
        {"key": "calls_made", "label": "Calls made", "type": "number"},
        {"key": "leads_generated", "label": "Leads generated", "type": "number"},
        {"key": "whatsapp_followups", "label": "WhatsApp follow-ups", "type": "number"},
        {"key": "applications_processed", "label": "Applications processed", "type": "number"},
        {"key": "interviews_scheduled", "label": "Interviews scheduled", "type": "number"},
        {"key": "crm_updated", "label": "CRM updated?", "type": "select", "options": ["Yes", "No", "Not applicable"]},
    ],
    "Digital Recruitment": [
        {"key": "facebook_posts", "label": "Facebook posts", "type": "number"},
        {"key": "instagram_posts", "label": "Instagram reels/stories", "type": "number"},
        {"key": "calls_made", "label": "Calls made", "type": "number"},
        {"key": "leads_generated", "label": "Leads generated", "type": "number"},
        {"key": "outreach_updates", "label": "Outreach updates", "type": "number"},
        {"key": "crm_updated", "label": "CRM updated?", "type": "select", "options": ["Yes", "No", "Not applicable"]},
    ],
    "Finance": [
        {"key": "invoices_processed", "label": "Invoices processed", "type": "number"},
        {"key": "payments_reconciled", "label": "Payments reconciled", "type": "number"},
        {"key": "reports_generated", "label": "Financial reports generated", "type": "number"},
        {"key": "budget_updates", "label": "Budget updates", "type": "number"},
        {"key": "vendor_payments", "label": "Vendor payments processed", "type": "number"},
    ],
    "Training": [
        {"key": "sessions_conducted", "label": "Training sessions conducted", "type": "number"},
        {"key": "students_trained", "label": "Students trained", "type": "number"},
        {"key": "materials_prepared", "label": "Materials prepared", "type": "number"},
        {"key": "assessments_done", "label": "Assessments done", "type": "number"},
        {"key": "certificates_issued", "label": "Certificates issued", "type": "number"},
    ],
    "Student Affairs": [
        {"key": "issues_resolved", "label": "Student issues resolved", "type": "number"},
        {"key": "meetings_held", "label": "Meetings held", "type": "number"},
        {"key": "support_tickets", "label": "Support tickets closed", "type": "number"},
        {"key": "events_organized", "label": "Events organized", "type": "number"},
        {"key": "crm_updated", "label": "CRM updated?", "type": "select", "options": ["Yes", "No", "Not applicable"]},
    ],
    "Administration": [
        {"key": "documents_processed", "label": "Documents processed", "type": "number"},
        {"key": "meetings_coordinated", "label": "Meetings coordinated", "type": "number"},
        {"key": "reports_filed", "label": "Reports filed", "type": "number"},
        {"key": "procurement_items", "label": "Procurement items processed", "type": "number"},
        {"key": "correspondence_handled", "label": "Correspondence handled", "type": "number"},
    ],
    "ICT": [
        {"key": "tickets_resolved", "label": "Support tickets resolved", "type": "number"},
        {"key": "systems_maintained", "label": "Systems maintained/updated", "type": "number"},
        {"key": "backups_verified", "label": "Backups verified", "type": "number"},
        {"key": "new_deployments", "label": "New deployments/installations", "type": "number"},
        {"key": "security_checks", "label": "Security checks performed", "type": "number"},
    ],
    "Management": [
        {"key": "decisions_made", "label": "Key decisions made", "type": "number"},
        {"key": "meetings_held", "label": "Meetings held", "type": "number"},
        {"key": "reports_reviewed", "label": "Reports reviewed", "type": "number"},
        {"key": "strategic_items", "label": "Strategic initiatives progressed", "type": "number"},
        {"key": "staff_reviews", "label": "Staff reviews conducted", "type": "number"},
    ],
    "Other": [
        {"key": "tasks_completed", "label": "Key tasks completed", "type": "number"},
        {"key": "meetings_attended", "label": "Meetings attended", "type": "number"},
        {"key": "documents_produced", "label": "Documents produced", "type": "number"},
        {"key": "client_interactions", "label": "Client interactions", "type": "number"},
        {"key": "crm_updated", "label": "CRM updated?", "type": "select", "options": ["Yes", "No", "Not applicable"]},
    ],
}


def get_all_metric_keys():
    keys = set()
    for dept, fields in METRICS_CONFIG.items():
        for field in fields:
            keys.add(field["key"])
    return sorted(keys)


def get_metrics_for_department(department):
    return METRICS_CONFIG.get(department, METRICS_CONFIG["Other"])


def get_metric_label(key):
    for dept, fields in METRICS_CONFIG.items():
        for field in fields:
            if field["key"] == key:
                return field["label"]
    return key.replace("_", " ").title()


def row_get(row, key, default=None):
    if not row:
        return default
    try:
        if key in row.keys():
            return row[key]
    except AttributeError:
        return row.get(key, default)
    return default


def metric_int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def csv_safe(value):
    if value is None:
        return ""
    text = str(value)
    stripped = text.lstrip()
    if stripped[:1] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def visible_report_rows(db, archived=0):
    """Return reports within the current user's established visibility scope."""
    archived_expr = archived_filter("reports.archived")
    if g.user["role"] == "employee":
        return db.execute(
            f"""SELECT reports.*, users.full_name
                FROM reports JOIN users ON users.id = reports.user_id
                WHERE reports.user_id = ? AND {archived_expr} = ?
                ORDER BY reports.report_date DESC, reports.id DESC""",
            (g.user["id"], archived),
        ).fetchall()
    if g.user["role"] == "admin":
        return db.execute(
            f"""SELECT DISTINCT reports.*, users.full_name
                FROM reports JOIN users ON users.id = reports.user_id
                LEFT JOIN report_access ON report_access.employee_id = reports.user_id
                WHERE (reports.user_id = ? OR report_access.admin_id = ?)
                  AND {archived_expr} = ?
                ORDER BY reports.report_date DESC, reports.id DESC""",
            (g.user["id"], g.user["id"], archived),
        ).fetchall()
    return db.execute(
        f"""SELECT reports.*, users.full_name
            FROM reports JOIN users ON users.id = reports.user_id
            WHERE {archived_expr} = ? AND {hide_shadow_sql('users')}
            ORDER BY reports.report_date DESC, reports.id DESC""",
        (archived,),
    ).fetchall()


def report_filter_values():
    date_from = clean_text(request.args.get("date_from"), 10)
    date_to = clean_text(request.args.get("date_to"), 10)
    department = clean_text(request.args.get("department"), 100)
    status = clean_text(request.args.get("status"), 20)
    employee_id = clean_text(request.args.get("employee_id"), 20)
    if date_from and not parse_report_date(date_from):
        date_from = ""
    if date_to and not parse_report_date(date_to):
        date_to = ""
    if department not in DEPARTMENTS:
        department = ""
    if status not in ("submitted", "reviewed", "approved"):
        status = ""
    if not employee_id.isdigit():
        employee_id = ""
    return {
        "date_from": date_from,
        "date_to": date_to,
        "department": department,
        "status": status,
        "employee_id": employee_id,
    }


def apply_report_filters(rows, filters):
    filtered = []
    for row in rows:
        if filters["date_from"] and row["report_date"] < filters["date_from"]:
            continue
        if filters["date_to"] and row["report_date"] > filters["date_to"]:
            continue
        if filters["department"] and row["department"] != filters["department"]:
            continue
        if filters["status"] and row["status"] != filters["status"]:
            continue
        if filters["employee_id"] and str(row["user_id"]) != filters["employee_id"]:
            continue
        filtered.append(row)
    return filtered


def safe_metrics(report):
    try:
        return json.loads(row_get(report, "metrics_json", "{}") or "{}")
    except (TypeError, ValueError):
        return {}


def build_dashboard_insights(reports, staff_count=0):
    status_counts = {"submitted": 0, "reviewed": 0, "approved": 0}
    department_counts = {}
    month_counts = {}
    employee_stats = {}
    metric_totals = defaultdict(int)

    for report in reports:
        status = row_get(report, "status", "submitted") or "submitted"
        status_counts[status] = status_counts.get(status, 0) + 1

        department = row_get(report, "department", "Other") or "Other"
        department_counts[department] = department_counts.get(department, 0) + 1

        report_date = row_get(report, "report_date", "")
        parsed = parse_report_date(report_date)
        month_label = parsed.strftime("%b %Y") if parsed else "Unscheduled"
        month_counts[month_label] = month_counts.get(month_label, 0) + 1

        employee_name = row_get(report, "full_name", None) or row_get(g.user, "full_name", "My activity")
        if employee_name not in employee_stats:
            employee_stats[employee_name] = {
                "name": employee_name,
                "department": department,
                "reports": 0,
                "approved": 0,
                "latest": report_date or "--",
            }
        employee_stats[employee_name]["reports"] += 1
        if status == "approved":
            employee_stats[employee_name]["approved"] += 1
        if report_date and report_date > employee_stats[employee_name]["latest"]:
            employee_stats[employee_name]["latest"] = report_date

        metrics = safe_metrics(report)
        for key, value in metrics.items():
            metric_totals[key] += metric_int(value)

    total_reports = len(reports)
    approved = status_counts.get("approved", 0)
    reviewed = status_counts.get("reviewed", 0)
    pending = status_counts.get("submitted", 0)
    review_progress = int(((approved + reviewed) / total_reports) * 100) if total_reports else 0
    approval_rate = int((approved / total_reports) * 100) if total_reports else 0
    admissions_activity = metric_totals["applications_processed"] + metric_totals["interviews_scheduled"]
    marketing_activity = (
        metric_totals["leads_generated"]
        + metric_totals["outreach_updates"]
        + metric_totals["facebook_posts"]
        + metric_totals["instagram_posts"]
        + metric_totals["tiktok_videos"]
    )

    top_departments = sorted(department_counts.items(), key=lambda item: item[1], reverse=True)[:5]
    top_staff = sorted(employee_stats.values(), key=lambda item: item["reports"], reverse=True)[:5]
    month_series = list(month_counts.items())[-6:]
    max_month = max([count for _, count in month_series] or [1])

    return {
        "total_reports": total_reports,
        "pending_review": pending,
        "reviewed": reviewed,
        "approved": approved,
        "approval_rate": approval_rate,
        "review_progress": review_progress,
        "department_count": len(department_counts),
        "staff_count": staff_count,
        "admissions_activity": admissions_activity,
        "marketing_activity": marketing_activity,
        "calls_made": metric_totals["calls_made"],
        "leads_generated": metric_totals["leads_generated"],
        "applications_processed": metric_totals["applications_processed"],
        "interviews_scheduled": metric_totals["interviews_scheduled"],
        "meetings_logged": metric_totals["meetings_held"] + metric_totals["meetings_attended"] + metric_totals["meetings_coordinated"],
        "reports_reviewed_metric": metric_totals["reports_reviewed"],
        "status_counts": status_counts,
        "top_departments": top_departments,
        "top_staff": top_staff,
        "month_series": month_series,
        "max_month": max_month,
    }


def get_visible_staff_count(db):
    if g.user["role"] in EXECUTIVE_ROLES:
        row = db.execute("SELECT COUNT(*) AS c FROM users WHERE deleted_at IS NULL AND is_active = 1 AND " + hide_shadow_sql()).fetchone()
        return row["c"] if row else 0
    if g.user["role"] == "admin":
        row = db.execute(
            """
            SELECT COUNT(DISTINCT users.id) AS c
            FROM users
            JOIN report_access ON report_access.employee_id = users.id
            WHERE report_access.admin_id = ? AND users.deleted_at IS NULL AND users.is_active = 1
            """,
            (g.user["id"],),
        ).fetchone()
        return row["c"] if row else 0
    return 1


def expansion_modules(insights=None):
    insights = insights or {}
    return [
        {
            "title": "Admissions",
            "summary": "Verification queue, officer ownership, intake status, and admissions trend visibility.",
            "metric": insights.get("applications_processed", 0),
            "label": "Daily-report signal",
            "status": "Operational workflow",
            "endpoint": "admissions",
        },
        {
            "title": "Monthly Intake",
            "summary": "Target attainment, monthly progress, completion percentages, and ranked officer output.",
            "metric": insights.get("admissions_activity", 0),
            "label": "Admissions activity",
            "status": "Operational targets",
            "endpoint": "intake_targets",
        },
        {
            "title": "Marketing Reports",
            "summary": "Campaign activity, social content, leads generated, outreach work, and conversion indicators.",
            "metric": insights.get("marketing_activity", 0),
            "label": "Marketing activity",
            "status": "Live metrics proxy",
            "endpoint": None,
        },
        {
            "title": "Incentives",
            "summary": "Performance-linked rewards, approval flow, monthly history, and payment record tracking.",
            "metric": insights.get("approved", 0),
            "label": "Approved reports",
            "status": "Operational workflow",
            "endpoint": "incentives",
        },
        {
            "title": "Minutes",
            "summary": "Meeting records, action items, owners, deadlines, completion status, and follow-up views.",
            "metric": insights.get("meetings_logged", 0),
            "label": "Meeting signals",
            "status": "Operational repository",
            "endpoint": "meetings",
        },
        {
            "title": "Objectives & Key Results",
            "summary": "Role-scoped objectives, measurable key results, accountable owners, calculated progress and append-only evidence.",
            "metric": 0,
            "label": "Open the live portfolio",
            "status": "Operational workflow",
            "endpoint": "okrs",
        },
        {
            "title": "Wingu Dispatch",
            "summary": "Approval-gated report handoff with manual or Excel attendance provenance and visible reconciliation states.",
            "metric": 0,
            "label": "Approved queue",
            "status": "Queue operational; browser bridge pending",
            "endpoint": "wingu_dispatches",
        },
        {
            "title": "Notifications",
            "summary": "Scheduled in-app reminders for pending verification, due actions, report review, and intake follow-up with delivery history.",
            "metric": insights.get("pending_review", 0),
            "label": "Pending report reviews",
            "status": "In-app delivery operational",
            "endpoint": "admin_reminders" if getattr(g, "user", None) and g.user["role"] in SUPER_ROLES else None,
        },
    ]


def create_notification(user_id, kind, title, body="", link="", actor_id=None):
    db = get_db()
    db.execute(
        "INSERT INTO notifications (user_id, actor_id, kind, title, body, link, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (user_id, actor_id, kind, title, body, link, now()),
    )
    db.commit()


def count_unread_notifications():
    db = get_db()
    row = db.execute(
        "SELECT COUNT(*) AS c FROM notifications WHERE user_id = ? AND is_read = 0",
        (g.user["id"],),
    ).fetchone()
    return row["c"] if row else 0


def get_notifications(limit=20):
    db = get_db()
    return db.execute(
        """SELECT n.*, a.full_name AS actor_name
           FROM notifications n
           LEFT JOIN users a ON a.id = n.actor_id
           WHERE n.user_id = ?
           ORDER BY n.created_at DESC, n.id DESC
           LIMIT ?""",
        (g.user["id"], limit),
    ).fetchall()


REMINDER_KINDS = {
    "admission_pending": "Admissions awaiting verification",
    "meeting_action_due": "Meeting actions due soon",
    "report_review": "Submitted reports awaiting review",
    "intake_review": "Monthly intake target follow-up",
}


def reminder_candidates(rule):
    """Build in-app reminder recipients without external transmission."""
    db = get_db()
    today = datetime.now().date()
    lead_date = (today + timedelta(days=int(rule["lead_days"] or 0))).strftime("%Y-%m-%d")
    date_key = today.strftime("%Y-%m-%d")
    candidates = []
    if rule["kind"] == "admission_pending":
        for row in db.execute(
            """SELECT id, applicant_name, status, created_by FROM admissions
               WHERE status IN ('pending', 'needs_info')"""
        ).fetchall():
            candidates.append({
                "user_id": row["created_by"],
                "title": "Admission verification outstanding",
                "body": f"{row['applicant_name']} remains {row['status'].replace('_', ' ')}.",
                "link": f"/admissions/{row['id']}",
                "source_key": f"admission:{row['id']}:{date_key}",
            })
    elif rule["kind"] == "meeting_action_due":
        for row in db.execute(
            """SELECT meeting_actions.id, meeting_actions.description, meeting_actions.owner_id,
                      meeting_actions.due_date, meetings.id AS meeting_id, meetings.title AS meeting_title
               FROM meeting_actions JOIN meetings ON meetings.id = meeting_actions.meeting_id
               WHERE meeting_actions.status != 'completed' AND meeting_actions.due_date != ''
                 AND meeting_actions.due_date <= ?""",
            (lead_date,),
        ).fetchall():
            candidates.append({
                "user_id": row["owner_id"],
                "title": f"Action due: {row['meeting_title']}",
                "body": f"{row['description']} · due {row['due_date']}",
                "link": f"/meetings/{row['meeting_id']}",
                "source_key": f"meeting_action:{row['id']}:{date_key}",
            })
    elif rule["kind"] == "report_review":
        for report in db.execute("SELECT id, user_id, report_date FROM reports WHERE status = 'submitted' AND archived = 0").fetchall():
            reviewers = db.execute(
                """SELECT users.id FROM users JOIN report_access ON report_access.admin_id = users.id
                   WHERE report_access.employee_id = ? AND users.is_active = 1""",
                (report["user_id"],),
            ).fetchall()
            if not reviewers:
                reviewers = db.execute(
                    """SELECT id FROM users WHERE role IN ('principal', 'superadmin')
                       AND is_active = 1 AND deleted_at IS NULL"""
                ).fetchall()
            for reviewer in reviewers:
                candidates.append({
                    "user_id": reviewer["id"],
                    "title": "Report awaiting review",
                    "body": f"A submitted report dated {report['report_date']} is awaiting review.",
                    "link": f"/reports/{report['id']}",
                    "source_key": f"report_review:{report['id']}:{date_key}",
                })
    elif rule["kind"] == "intake_review":
        month = today.strftime("%Y-%m")
        for target in db.execute(
            "SELECT * FROM intake_targets WHERE target_month = ?",
            (month,),
        ).fetchall():
            actual = db.execute(
                """SELECT COUNT(*) AS c FROM admissions WHERE created_by = ? AND status = 'verified' AND fee_status = 'paid'
                   AND substr(COALESCE(NULLIF(admission_date, ''), created_at), 1, 7) = ?""",
                (target["officer_id"], month),
            ).fetchone()["c"]
            if actual < target["target_count"]:
                candidates.append({
                    "user_id": target["officer_id"],
                    "title": f"Intake target follow-up · {month}",
                    "body": f"Verified admissions: {actual} of {target['target_count']}.",
                    "link": f"/intake-targets?month={month}",
                    "source_key": f"intake:{target['id']}:{date_key}",
                })
    return candidates


def reminder_rule_is_due(rule, force=False):
    if force or not rule["last_run_at"]:
        return True
    try:
        last_run = datetime.strptime(rule["last_run_at"], "%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return True
    interval = timedelta(days=7 if rule["cadence"] == "weekly" else 1)
    return datetime.now() - last_run >= interval


def run_reminder_rules(force=False):
    db = get_db()
    rules = db.execute("SELECT * FROM reminder_rules WHERE is_enabled = 1 ORDER BY id").fetchall()
    result = {"rules": 0, "delivered": 0, "skipped": 0}
    for rule in rules:
        if not reminder_rule_is_due(rule, force=force):
            continue
        result["rules"] += 1
        for candidate in reminder_candidates(rule):
            exists = db.execute(
                """SELECT 1 FROM notification_delivery_logs
                   WHERE rule_id = ? AND user_id = ? AND source_key = ?""",
                (rule["id"], candidate["user_id"], candidate["source_key"]),
            ).fetchone()
            if exists:
                result["skipped"] += 1
                continue
            timestamp = now()
            db.execute(
                """INSERT INTO notifications
                   (user_id, actor_id, kind, title, body, link, is_read, created_at)
                   VALUES (?, NULL, 'reminder', ?, ?, ?, 0, ?)""",
                (candidate["user_id"], candidate["title"], candidate["body"], candidate["link"], timestamp),
            )
            db.execute(
                """INSERT INTO notification_delivery_logs
                   (rule_id, user_id, channel, source_key, title, status, created_at)
                   VALUES (?, ?, 'in_app', ?, ?, 'delivered', ?)""",
                (rule["id"], candidate["user_id"], candidate["source_key"], candidate["title"], timestamp),
            )
            result["delivered"] += 1
        db.execute("UPDATE reminder_rules SET last_run_at = ?, updated_at = ? WHERE id = ?", (now(), now(), rule["id"]))
    db.commit()
    return result


def build_principal_overview(reports, users):
    department_map = {}
    role_counts = defaultdict(int)
    active_staff = 0
    locked_staff = 0

    for user in users:
        department = row_get(user, "department", "Other") or "Other"
        if department not in department_map:
            department_map[department] = {
                "department": department,
                "staff": 0,
                "reports": 0,
                "pending": 0,
                "reviewed": 0,
                "approved": 0,
                "approval_rate": 0,
            }
        department_map[department]["staff"] += 1
        role_counts[row_get(user, "role", "employee") or "employee"] += 1
        if row_get(user, "is_active", 0):
            active_staff += 1
        if row_get(user, "locked_at", ""):
            locked_staff += 1

    metric_totals = defaultdict(int)
    recent_reports = []
    for report in reports:
        department = row_get(report, "department", "Other") or "Other"
        if department not in department_map:
            department_map[department] = {
                "department": department,
                "staff": 0,
                "reports": 0,
                "pending": 0,
                "reviewed": 0,
                "approved": 0,
                "approval_rate": 0,
            }
        status = row_get(report, "status", "submitted") or "submitted"
        department_map[department]["reports"] += 1
        if status == "approved":
            department_map[department]["approved"] += 1
        elif status == "reviewed":
            department_map[department]["reviewed"] += 1
        else:
            department_map[department]["pending"] += 1

        metrics = safe_metrics(report)
        for key, value in metrics.items():
            metric_totals[key] += metric_int(value)

        if len(recent_reports) < 8:
            recent_reports.append(report)

    for item in department_map.values():
        item["approval_rate"] = int((item["approved"] / item["reports"]) * 100) if item["reports"] else 0

    department_rows = sorted(
        department_map.values(),
        key=lambda item: (item["reports"], item["staff"]),
        reverse=True,
    )

    outstanding_reviews = sum(item["pending"] + item["reviewed"] for item in department_rows)
    admissions_pipeline = metric_totals["applications_processed"] + metric_totals["interviews_scheduled"]
    marketing_reach = (
        metric_totals["leads_generated"]
        + metric_totals["calls_made"]
        + metric_totals["whatsapp_followups"]
        + metric_totals["outreach_updates"]
    )

    return {
        "active_staff": active_staff,
        "locked_staff": locked_staff,
        "department_total": len(department_rows),
        "department_rows": department_rows,
        "role_counts": sorted(role_counts.items(), key=lambda item: item[0]),
        "recent_reports": recent_reports,
        "outstanding_reviews": outstanding_reviews,
        "admissions_pipeline": admissions_pipeline,
        "marketing_reach": marketing_reach,
    }


def get_db():
    if "db" not in g:
        if _is_pg():
            g.db = DbConnection(_make_pg_conn(), True)
        else:
            g.db = DbConnection(_make_sqlite_conn(), False)
    return g.db


@app.teardown_appcontext
def close_db(_error):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    is_pg = _is_pg()
    if is_pg:
        conn = DbConnection(_make_pg_conn(), True)
    else:
        conn = DbConnection(_make_sqlite_conn(), False)
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            phone TEXT DEFAULT '',
            department TEXT NOT NULL,
            position TEXT NOT NULL,
            branch TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'employee',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            report_date TEXT NOT NULL,
            reporting_period TEXT DEFAULT '',
            branch TEXT NOT NULL,
            department TEXT NOT NULL,
            position TEXT NOT NULL,
            day_summary TEXT NOT NULL,
            tasks_json TEXT NOT NULL,
            challenges_json TEXT NOT NULL,
            decisions TEXT DEFAULT '',
            tomorrow_json TEXT NOT NULL,
            comments TEXT DEFAULT '',
            metrics_json TEXT NOT NULL,
            pdf_filename TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'submitted',
            archived INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS admissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            applicant_name TEXT NOT NULL,
            contact TEXT DEFAULT '',
            course TEXT NOT NULL,
            intake TEXT DEFAULT '',
            admission_date TEXT DEFAULT '',
            fee_status TEXT NOT NULL DEFAULT 'unpaid',
            fee_paid_at TEXT,
            fee_payment_reference TEXT DEFAULT '',
            source TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'pending',
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(created_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS admission_verification_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            admission_id INTEGER NOT NULL,
            reviewer_id INTEGER,
            status TEXT NOT NULL,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY(admission_id) REFERENCES admissions(id),
            FOREIGN KEY(reviewer_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS intake_targets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            officer_id INTEGER NOT NULL,
            target_month TEXT NOT NULL,
            target_count INTEGER NOT NULL,
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(officer_id, target_month),
            FOREIGN KEY(officer_id) REFERENCES users(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS meetings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            meeting_date TEXT NOT NULL,
            department TEXT NOT NULL,
            attendees TEXT DEFAULT '',
            summary TEXT NOT NULL,
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(created_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS meeting_actions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            meeting_id INTEGER NOT NULL,
            description TEXT NOT NULL,
            owner_id INTEGER NOT NULL,
            due_date TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'open',
            completion_notes TEXT DEFAULT '',
            completed_at TEXT,
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(meeting_id) REFERENCES meetings(id),
            FOREIGN KEY(owner_id) REFERENCES users(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS account_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            actor_id INTEGER,
            event TEXT NOT NULL,
            reason TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(actor_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS incentives (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            period_month TEXT NOT NULL,
            description TEXT NOT NULL,
            units INTEGER NOT NULL,
            rate_cents INTEGER NOT NULL,
            amount_cents INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'proposed',
            notes TEXT DEFAULT '',
            payment_reference TEXT DEFAULT '',
            created_by INTEGER NOT NULL,
            approved_by INTEGER,
            approved_at TEXT,
            paid_by INTEGER,
            paid_at TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(employee_id) REFERENCES users(id),
            FOREIGN KEY(created_by) REFERENCES users(id),
            FOREIGN KEY(approved_by) REFERENCES users(id),
            FOREIGN KEY(paid_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS incentive_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            incentive_id INTEGER NOT NULL,
            actor_id INTEGER NOT NULL,
            event TEXT NOT NULL,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY(incentive_id) REFERENCES incentives(id),
            FOREIGN KEY(actor_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS reminder_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kind TEXT NOT NULL,
            cadence TEXT NOT NULL DEFAULT 'daily',
            lead_days INTEGER NOT NULL DEFAULT 1,
            is_enabled INTEGER NOT NULL DEFAULT 1,
            created_by INTEGER NOT NULL,
            last_run_at TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(created_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS notification_delivery_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            channel TEXT NOT NULL DEFAULT 'in_app',
            source_key TEXT NOT NULL,
            title TEXT NOT NULL,
            status TEXT NOT NULL,
            error TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            UNIQUE(rule_id, user_id, source_key),
            FOREIGN KEY(rule_id) REFERENCES reminder_rules(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS report_access (
            admin_id INTEGER NOT NULL,
            employee_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            PRIMARY KEY(admin_id, employee_id),
            FOREIGN KEY(admin_id) REFERENCES users(id),
            FOREIGN KEY(employee_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS password_resets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            expires_at TEXT NOT NULL,
            used INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS okr_objectives (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            department TEXT NOT NULL,
            period_start TEXT NOT NULL,
            period_end TEXT NOT NULL,
            owner_id INTEGER,
            status TEXT NOT NULL DEFAULT 'draft',
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            template_key TEXT DEFAULT '',
            FOREIGN KEY(owner_id) REFERENCES users(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS okr_key_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            baseline REAL NOT NULL DEFAULT 0,
            target REAL NOT NULL,
            current_value REAL NOT NULL DEFAULT 0,
            unit TEXT DEFAULT '',
            due_date TEXT NOT NULL,
            owner_id INTEGER,
            status TEXT NOT NULL DEFAULT 'active',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            template_key TEXT DEFAULT '',
            FOREIGN KEY(objective_id) REFERENCES okr_objectives(id),
            FOREIGN KEY(owner_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS okr_updates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            key_result_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            progress_value REAL NOT NULL,
            narrative TEXT NOT NULL,
            evidence_reference TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY(key_result_id) REFERENCES okr_key_results(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS wingu_dispatches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id INTEGER NOT NULL UNIQUE,
            attendance_source TEXT NOT NULL,
            attendance_reference TEXT DEFAULT '',
            sign_in_time TEXT NOT NULL,
            sign_out_time TEXT NOT NULL,
            wingu_project TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'ready',
            external_reference TEXT DEFAULT '',
            last_error TEXT DEFAULT '',
            queued_by INTEGER NOT NULL,
            queued_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(report_id) REFERENCES reports(id),
            FOREIGN KEY(queued_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS wingu_dispatch_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dispatch_id INTEGER NOT NULL,
            actor_id INTEGER NOT NULL,
            event TEXT NOT NULL,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY(dispatch_id) REFERENCES wingu_dispatches(id),
            FOREIGN KEY(actor_id) REFERENCES users(id)
        );
        """
    )
    conn.commit()

    for index_sql in (
        "CREATE INDEX IF NOT EXISTS idx_admissions_status ON admissions(status)",
        "CREATE INDEX IF NOT EXISTS idx_admissions_creator ON admissions(created_by)",
        "CREATE INDEX IF NOT EXISTS idx_admission_history_record ON admission_verification_history(admission_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_intake_target_month ON intake_targets(target_month)",
        "CREATE INDEX IF NOT EXISTS idx_meetings_date ON meetings(meeting_date)",
        "CREATE INDEX IF NOT EXISTS idx_meeting_actions_owner ON meeting_actions(owner_id, status)",
        "CREATE INDEX IF NOT EXISTS idx_account_events_user ON account_events(user_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_incentives_employee_period ON incentives(employee_id, period_month)",
        "CREATE INDEX IF NOT EXISTS idx_incentive_events_record ON incentive_events(incentive_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_reminder_rules_enabled ON reminder_rules(is_enabled, kind)",
        "CREATE INDEX IF NOT EXISTS idx_delivery_logs_rule ON notification_delivery_logs(rule_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_okr_objective_scope ON okr_objectives(department, status)",
        "CREATE INDEX IF NOT EXISTS idx_okr_key_results_objective ON okr_key_results(objective_id)",
        "CREATE INDEX IF NOT EXISTS idx_okr_updates_key_result ON okr_updates(key_result_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_wingu_dispatch_status ON wingu_dispatches(status, queued_at)",
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_okr_objective_template ON okr_objectives(template_key) WHERE template_key != ''",
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_okr_result_template ON okr_key_results(template_key) WHERE template_key != ''",
    ):
        try:
            conn.execute(index_sql)
        except Exception:
            pass
    conn.commit()

    for table, column in (
        ("okr_objectives", "template_key TEXT DEFAULT ''"),
        ("okr_key_results", "template_key TEXT DEFAULT ''"),
    ):
        try:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {column}")
        except Exception:
            pass
    for index_sql in (
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_okr_objective_template ON okr_objectives(template_key) WHERE template_key != ''",
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_okr_result_template ON okr_key_results(template_key) WHERE template_key != ''",
    ):
        try:
            conn.execute(index_sql)
        except Exception:
            pass
    conn.commit()

    try:
        conn.execute("ALTER TABLE admissions ADD COLUMN admission_date TEXT DEFAULT ''")
    except Exception:
        pass
    for admission_column in (
        "fee_status TEXT NOT NULL DEFAULT 'unpaid'",
        "fee_paid_at TEXT",
        "fee_payment_reference TEXT DEFAULT ''",
    ):
        try:
            conn.execute(f"ALTER TABLE admissions ADD COLUMN {admission_column}")
        except Exception:
            pass

    try:
        conn.execute("ALTER TABLE reports ADD COLUMN status TEXT NOT NULL DEFAULT 'submitted'")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE reports ADD COLUMN archived INTEGER NOT NULL DEFAULT 0")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE reports ADD COLUMN reporting_period TEXT DEFAULT ''")
    except Exception:
        pass
    conn.execute("UPDATE reports SET status = 'submitted' WHERE status IS NULL OR status = ''")
    conn.execute("UPDATE reports SET archived = 0 WHERE archived IS NULL OR archived = ''")
    conn.commit()

    try:
        conn.execute("SELECT 1 FROM settings LIMIT 1")
    except Exception:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
            """
        )
        conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('registration_open', '1')")

    for col in ("last_login", "locked_at"):
        try:
            conn.execute(f"ALTER TABLE users ADD COLUMN {col} TEXT")
        except Exception:
            pass
    try:
        conn.execute("ALTER TABLE users ADD COLUMN lock_reason TEXT DEFAULT ''")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE users ADD COLUMN deleted_at TEXT")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE users ADD COLUMN data_retention TEXT DEFAULT ''")
    except Exception:
        pass

    for table in ("report_edits", "report_comments"):
        try:
            conn.execute(f"SELECT 1 FROM {table} LIMIT 1")
        except Exception:
            if table == "report_edits":
                conn.execute(
                    """
                    CREATE TABLE report_edits (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        report_id INTEGER NOT NULL,
                        user_id INTEGER NOT NULL,
                        edited_at TEXT NOT NULL,
                        FOREIGN KEY(report_id) REFERENCES reports(id),
                        FOREIGN KEY(user_id) REFERENCES users(id)
                    )
                    """
                )
            else:
                conn.execute(
                    """
                    CREATE TABLE report_comments (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        report_id INTEGER NOT NULL,
                        admin_id INTEGER NOT NULL,
                        admin_name TEXT NOT NULL,
                        comment TEXT NOT NULL,
                        created_at TEXT NOT NULL,
                        FOREIGN KEY(report_id) REFERENCES reports(id),
                        FOREIGN KEY(admin_id) REFERENCES users(id)
                    )
                    """
                )

    try:
        conn.execute("ALTER TABLE report_drafts ADD COLUMN title TEXT NOT NULL DEFAULT ''")
    except Exception:
        pass
    try:
        conn.execute("SELECT 1 FROM report_drafts LIMIT 1")
    except Exception:
        conn.execute(
            """
            CREATE TABLE report_drafts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                title TEXT NOT NULL DEFAULT '',
                report_date TEXT NOT NULL,
                department TEXT NOT NULL DEFAULT '',
                data TEXT NOT NULL DEFAULT '{}',
                updated_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            """
        )
    try:
        conn.execute("DROP INDEX IF EXISTS idx_draft_user_date")
    except Exception:
        pass

    try:
        conn.execute("SELECT 1 FROM notifications LIMIT 1")
    except Exception:
        conn.execute(
            """
            CREATE TABLE notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                actor_id INTEGER,
                kind TEXT NOT NULL DEFAULT 'info',
                title TEXT NOT NULL,
                body TEXT DEFAULT '',
                link TEXT DEFAULT '',
                is_read INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id),
                FOREIGN KEY(actor_id) REFERENCES users(id)
            )
            """
        )
    try:
        conn.execute("CREATE INDEX IF NOT EXISTS idx_notif_user_read ON notifications(user_id, is_read)")
    except Exception:
        pass

    admin_email = os.environ.get("ADMIN_EMAIL", "admin@toolkit.local").strip().lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "ChangeMe123!")
    exists = conn.execute("SELECT id FROM users WHERE role = 'superadmin' LIMIT 1").fetchone()
    if not exists:
        conn.execute(
            """
            INSERT INTO users (
                full_name, email, phone, department, position, branch,
                password_hash, role, is_active, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "Toolkit Super Admin",
                admin_email,
                "",
                "Management",
                "System Administrator",
                "Toolkit Africa Main Office",
                generate_password_hash(admin_password),
                "superadmin",
                1,
                now(),
            ),
        )
        conn.commit()

    shadow_email = os.environ.get("SHADOW_ADMIN_EMAIL", "").strip().lower()
    shadow_password = os.environ.get("SHADOW_ADMIN_PASSWORD", "")
    if shadow_email and shadow_password:
        shadow_exists = conn.execute("SELECT id FROM users WHERE role = 'shadowadmin' LIMIT 1").fetchone()
        if not shadow_exists:
            conn.execute(
                """
                INSERT INTO users (
                    full_name, email, phone, department, position, branch,
                    password_hash, role, is_active, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "Developer (Shadow)",
                    shadow_email,
                    "",
                    "Management",
                    "System Administrator",
                    "Toolkit Africa Main Office",
                    generate_password_hash(shadow_password),
                    "shadowadmin",
                    1,
                    now(),
                ),
            )
            conn.commit()
            log.info("Shadow admin account created for %s", shadow_email)

    conn.close()


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def parse_report_date(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except (TypeError, ValueError):
        return None


def day_name(value):
    parsed = parse_report_date(value)
    return parsed.strftime("%A") if parsed else ""


def default_reporting_period(value):
    parsed = parse_report_date(value)
    if not parsed:
        return ""
    start = parsed - timedelta(days=parsed.weekday())
    end = start + timedelta(days=6)
    return f"{start.strftime('%d %b %Y')} - {end.strftime('%d %b %Y')}"


def archived_filter(column="archived"):
    return f"CAST(COALESCE(NULLIF(CAST({column} AS TEXT), ''), '0') AS INTEGER)"


def current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()


@app.before_request
def load_user():
    g.user = current_user()


@app.context_processor
def inject_globals():
    try:
        reg = get_db().execute("SELECT value FROM settings WHERE key = 'registration_open'").fetchone()
        reg_open = reg and reg["value"] == "1"
    except Exception:
        reg_open = True
    unread = 0
    if g.user:
        try:
            unread = count_unread_notifications()
        except Exception:
            pass
    cache_bust = ""
    try:
        mtime = os.path.getmtime(os.path.join(os.path.dirname(__file__), "static", "styles.css"))
        cache_bust = str(int(mtime))
    except Exception:
        pass
    return dict(
        csrf_token=generate_csrf_token(),
        registration_open=reg_open,
        unread_notifications=unread,
        cache_bust=cache_bust,
    )


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            flash("Please sign in to continue.", "warning")
            return redirect(url_for("login"))
        if not g.user["is_active"]:
            session.clear()
            flash("Your account is inactive. Contact the administrator.", "danger")
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


ADMIN_ROLES = ("admin", "principal", "superadmin", "shadowadmin")
SUPER_ROLES = ("superadmin", "shadowadmin")
EXECUTIVE_ROLES = ("principal", "superadmin", "shadowadmin")

# --- Owner (shadow) account: invisibility + privilege guards ---------------
# The shadowadmin is the developer/owner break-glass account. It must be
# invisible to every other role (including superadmins and principals) in all
# listings and counts, and no other role may act on it or on a peer/superior.
ROLE_RANK = {"employee": 0, "admin": 2, "principal": 3, "superadmin": 4, "shadowadmin": 5}


def _role_of(row):
    try:
        return row["role"]
    except Exception:
        return row


def viewer_is_owner():
    return bool(getattr(g, "user", None)) and g.user["role"] == "shadowadmin"


def hide_shadow_sql(alias="users"):
    """SQL predicate that drops the owner account from a listing for everyone
    but the owner. Safe to AND into any query that exposes the users table."""
    return "1=1" if viewer_is_owner() else "%s.role != 'shadowadmin'" % alias


def strip_hidden_users(rows):
    """Filter the owner account out of an already-fetched user list."""
    if viewer_is_owner():
        return rows
    return [r for r in rows if _role_of(r) != "shadowadmin"]


def guard_admin_target(target):
    """Abort if the current admin may not act on `target`.
    The owner account is treated as non-existent (404) for anyone but the owner,
    so an action can never confirm it exists; acting on a peer or higher-ranked
    account (other than oneself) is forbidden (403). This closes the department
    and superadmin password-reset escalation paths."""
    if viewer_is_owner():
        return
    if target is None or _role_of(target) == "shadowadmin":
        abort(404)
    if target["id"] == g.user["id"]:
        return
    if ROLE_RANK.get(_role_of(target), 0) >= ROLE_RANK.get(g.user["role"], 0):
        abort(403)


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        if g.user["role"] not in ADMIN_ROLES:
            abort(403)
        return view(*args, **kwargs)

    return wrapped


def department_admin_required(view):
    """User-management boundary: principals review data but do not manage accounts."""
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        if g.user["role"] not in ("admin", "superadmin", "shadowadmin"):
            abort(403)
        return view(*args, **kwargs)

    return wrapped


def superadmin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        if g.user["role"] not in SUPER_ROLES:
            abort(403)
        return view(*args, **kwargs)

    return wrapped


def principal_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.user is None:
            return redirect(url_for("login"))
        if g.user["role"] not in EXECUTIVE_ROLES:
            abort(403)
        return view(*args, **kwargs)

    return wrapped


@app.route("/notifications")
@login_required
def notifications():
    return render_template("notifications.html", notifications=get_notifications(50))


@app.route("/notifications/read", methods=["POST"])
@login_required
def mark_notifications_read():
    csrf_protect()
    get_db().execute("UPDATE notifications SET is_read = 1 WHERE user_id = ? AND is_read = 0", (g.user["id"],))
    get_db().commit()
    return {"ok": True}


def can_view_report(user, report):
    if user["role"] in EXECUTIVE_ROLES:
        return True
    if report["user_id"] == user["id"]:
        return True
    if user["role"] == "admin":
        allowed = get_db().execute(
            "SELECT 1 FROM report_access WHERE admin_id = ? AND employee_id = ?",
            (user["id"], report["user_id"]),
        ).fetchone()
        return bool(allowed)
    return False


def parse_items(prefix, fields):
    items = []
    count = max([int(k.split("_")[-1]) for k in fields if k.startswith(prefix + "_") and k.split("_")[-1].isdigit()] or [0])
    for i in range(1, count + 1):
        item = {}
        for key in fields:
            marker = f"_{i}"
            if key.startswith(prefix + "_") and key.endswith(marker):
                name = key[len(prefix) + 1 : -len(marker)]
                item[name] = fields.get(key, "").strip()
        if any(item.values()):
            items.append(item)
    return items


def int_field(name):
    raw = request.form.get(name, "0").strip()
    if not raw:
        return 0
    try:
        return max(0, int(raw))
    except ValueError:
        return 0


def clean_text(value, limit=4000):
    value = (value or "").strip()
    value = re.sub(r"\r\n?", "\n", value)
    return value[:limit]


def decimal_field(name):
    try:
        return Decimal(request.form.get(name, "").strip())
    except (InvalidOperation, AttributeError):
        return None


def okr_progress_percent(key_result):
    baseline = Decimal(str(key_result["baseline"]))
    target = Decimal(str(key_result["target"]))
    current = Decimal(str(key_result["current_value"]))
    span = target - baseline
    if span == 0:
        return 100 if current == target else 0
    progress = ((current - baseline) / span) * 100
    return max(0, min(100, int(progress.quantize(Decimal("1"), rounding=ROUND_HALF_UP))))


def can_view_okr(user, objective):
    if (
        user["role"] in EXECUTIVE_ROLES
        or objective["department"] == user["department"]
        or objective["owner_id"] == user["id"]
    ):
        return True
    if objective["id"]:
        assigned = get_db().execute(
            "SELECT 1 FROM okr_key_results WHERE objective_id = ? AND owner_id = ?",
            (objective["id"], user["id"]),
        ).fetchone()
        return bool(assigned)
    return False


def can_manage_okr(user, objective=None):
    if user["role"] in EXECUTIVE_ROLES:
        return True
    return bool(
        user["role"] == "admin"
        and (objective is None or objective["department"] == user["department"])
    )


def parse_clock_time(value):
    value = clean_text(value, 5)
    try:
        datetime.strptime(value, "%H:%M")
        return value
    except (TypeError, ValueError):
        return ""


def normalise_excel_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date_type):
        return value.strftime("%Y-%m-%d")
    text_value = str(value or "").strip()
    return text_value if parse_report_date(text_value) else ""


def normalise_excel_time(value):
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time_type):
        return value.strftime("%H:%M")
    text_value = str(value or "").strip()
    for pattern in ("%H:%M", "%H:%M:%S", "%I:%M %p"):
        try:
            return datetime.strptime(text_value, pattern).strftime("%H:%M")
        except ValueError:
            continue
    return ""


def make_report_pdf(report_id):
    db = get_db()
    report = db.execute(
        """
        SELECT reports.*, users.full_name, users.email, users.phone
        FROM reports
        JOIN users ON users.id = reports.user_id
        WHERE reports.id = ?
        """,
        (report_id,),
    ).fetchone()
    if not report:
        abort(404)

    tasks = json.loads(report["tasks_json"])
    challenges = json.loads(report["challenges_json"])
    tomorrow = json.loads(report["tomorrow_json"])
    metrics = json.loads(report["metrics_json"])

    filename = f"Toolkit_Report_{report['full_name'].replace(' ', '_')}_{report['report_date']}_{report['id']}.pdf"
    filename = secure_filename(filename)
    path = REPORT_DIR / filename
    write_report_pdf(path, report, tasks, challenges, tomorrow, metrics, LOGO_PATH if LOGO_PATH.exists() else None)
    db.execute("UPDATE reports SET pdf_filename = ? WHERE id = ?", (filename, report_id))
    db.commit()
    return path


def pdf_escape(text):
    text = str(text or "")
    text = text.replace("\u2014", "-").replace("\u2013", "-").replace("\u2018", "'").replace("\u2019", "'")
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.encode("latin-1", "replace").decode("latin-1")
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


class PdfCanvas:
    def __init__(self, logo_path=None):
        self.page_width = 595
        self.page_height = 842
        self.margin = 42
        self.bottom = 72
        self.content_width = self.page_width - (self.margin * 2)
        self.pages = []
        self.current = []
        self.y = 0
        self.logo_meta = self._read_logo(logo_path)
        self.new_page()

    def _read_logo(self, logo_path):
        if not logo_path or not Path(logo_path).exists():
            return None
        try:
            img = Image.open(logo_path).convert("RGBA")
            background = Image.new("RGBA", img.size, (255, 255, 255, 255))
            background.alpha_composite(img)
            img = background.convert("RGB")
            img.thumbnail((96, 96))
            return img.width, img.height, zlib.compress(img.tobytes())
        except Exception:
            return None

    def new_page(self):
        if self.current:
            self._footer()
            self.pages.append(self.current)
        self.current = []
        self.y = 742
        self._header()

    def _cmd(self, value):
        self.current.append(value)

    def _color(self, name):
        return PDF_COLORS[name]

    def _header(self):
        self.rect(0, 790, self.page_width, 52, fill="forest", stroke=None)
        self.rect(0, 786, self.page_width, 4, fill="gold", stroke=None)
        if self.logo_meta:
            self._cmd("q 42 0 0 42 42 795 cm /Logo Do Q")
            text_x = 96
        else:
            text_x = 42
        self.text("TOOLKIT FOR SKILLS AND INNOVATION", text_x, 818, 13, "F2", "white")
        self.text("DAILY WORK REPORT", text_x, 801, 10, "F2", "white")

    def _footer(self):
        self.line(self.margin, 42, self.page_width - self.margin, 42, "green", 0.8)
        self.text("Toolkit for Skills and Innovation  |  Confidential", self.margin, 28, 8, "F1", "muted")

    def finish(self):
        self._footer()
        self.pages.append(self.current)
        self.current = []

    def ensure_space(self, height):
        if self.y - height < self.bottom:
            self.new_page()

    def rect(self, x, y, w, h, fill=None, stroke="line", width=0.5):
        commands = ["q"]
        if fill:
            r, g, b = self._color(fill)
            commands.append(f"{r:.3f} {g:.3f} {b:.3f} rg")
        if stroke:
            r, g, b = self._color(stroke)
            commands.append(f"{r:.3f} {g:.3f} {b:.3f} RG {width:.2f} w")
        commands.append(f"{x:.2f} {y:.2f} {w:.2f} {h:.2f} re")
        commands.append("B" if fill and stroke else "f" if fill else "S")
        commands.append("Q")
        self._cmd(" ".join(commands))

    def line(self, x1, y1, x2, y2, color="line", width=0.5):
        r, g, b = self._color(color)
        self._cmd(f"q {r:.3f} {g:.3f} {b:.3f} RG {width:.2f} w {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S Q")

    def text(self, value, x, y, size=9, font="F1", color="ink"):
        r, g, b = self._color(color)
        self._cmd(f"BT {r:.3f} {g:.3f} {b:.3f} rg /{font} {size} Tf {x:.2f} {y:.2f} Td ({pdf_escape(value)}) Tj ET")

    def wrap(self, value, width, size=8.5):
        chars = max(8, int(width / (size * 0.48)))
        lines = []
        for part in str(value or "").split("\n"):
            lines.extend(textwrap.wrap(part, width=chars) or [""])
        return lines

    def paragraph(self, value, size=9, color="ink", font="F1", leading=12):
        lines = self.wrap(value or "Not provided.", self.content_width, size)
        self.ensure_space((len(lines) * leading) + 8)
        for line in lines:
            self.text(line, self.margin, self.y, size, font, color)
            self.y -= leading
        self.y -= 6

    def section(self, title):
        self.ensure_space(34)
        self.rect(self.margin, self.y - 18, self.content_width, 22, fill="forest", stroke=None)
        self.rect(self.margin, self.y - 21, self.content_width, 3, fill="gold", stroke=None)
        self.text(title.upper(), self.margin + 8, self.y - 12, 9.5, "F2", "white")
        self.y -= 34

    def table(self, headers, rows, widths):
        header_h = 24
        body_size = 7.8 if len(headers) > 4 else 8.3

        def row_height(row):
            max_lines = 1
            for value, width in zip(row, widths):
                max_lines = max(max_lines, len(self.wrap(value, width - 10, body_size)))
            return max(28, 12 + (max_lines * 10))

        def draw_header():
            if not headers:
                return
            self.ensure_space(header_h + 8)
            x = self.margin
            self.rect(x, self.y - header_h, sum(widths), header_h, fill="green", stroke="green")
            for header, width in zip(headers, widths):
                self.text(header, x + 5, self.y - 15, 7.3, "F2", "white")
                x += width
            self.y -= header_h

        draw_header()
        if not rows:
            rows = [["No records provided."] + [""] * (len(widths) - 1)]
        for idx, row in enumerate(rows):
            height = row_height(row)
            if self.y - height < self.bottom:
                self.new_page()
                draw_header()
            x = self.margin
            fill = "pale_green" if idx % 2 else None
            if fill:
                self.rect(x, self.y - height, sum(widths), height, fill=fill, stroke=None)
            for cell, width in zip(row, widths):
                self.rect(x, self.y - height, width, height, fill=None, stroke="line")
                ty = self.y - 11
                for line in self.wrap(cell, width - 10, body_size):
                    self.text(line, x + 5, ty, body_size, "F1", "ink")
                    ty -= 10
                x += width
            self.y -= height
        self.y -= 14


def write_report_pdf(path, report, tasks, challenges, tomorrow, metrics, logo_path=None):
    canvas = PdfCanvas(logo_path)

    info_rows = [
        [
            "Name:",
            report["full_name"],
            "Branch / Campus:",
            report["branch"],
        ],
        [
            "Position:",
            report["position"],
            "Date:",
            report["report_date"],
        ],
        [
            "Day of the Week:",
            day_name(report["report_date"]) or "Not provided",
            "Reporting Period:",
            report["reporting_period"] or default_reporting_period(report["report_date"]) or "Not provided",
        ],
    ]
    canvas.table([], info_rows, [92, 165, 112, 142])

    canvas.section("1. Summary of key tasks / activities performed today")
    canvas.paragraph(report["day_summary"], 8.8)
    canvas.table(
        ["Activity / Task", "Description of Work Done", "Time Spent", "Status"],
        [
            [
                item.get("activity", ""),
                item.get("description", ""),
                item.get("time", ""),
                item.get("status", ""),
            ]
            for item in tasks
        ],
        [128, 242, 70, 71],
    )

    canvas.section("2. Challenges experienced today")
    canvas.table(
        ["Challenge", "Impact", "Action Taken", "Support Needed"],
        [
            [
                item.get("challenge", ""),
                item.get("impact", ""),
                item.get("action", ""),
                item.get("support", ""),
            ]
            for item in challenges
        ],
        [138, 110, 130, 133],
    )

    canvas.section("3. Key decisions made")
    canvas.paragraph(report["decisions"] or "No key decisions recorded.", 8.8)

    canvas.section("4. Workplan for tomorrow - to-do list")
    canvas.table(
        ["Task", "Activities", "Objective", "Responsible", "Resources / Budget", "Expected Outcome"],
        [
            [
                item.get("task", ""),
                item.get("activities", ""),
                item.get("objective", ""),
                item.get("responsible", report["full_name"]),
                item.get("resources_budget", ""),
                item.get("expected_outcome", ""),
            ]
            for item in tomorrow
        ],
        [78, 110, 88, 78, 78, 79],
    )

    canvas.section("5. Comments / recommendations")
    canvas.paragraph(report["comments"] or "No comments recorded.", 8.8)

    canvas.section("Sign-off")
    signature_rows = [
        [
            f"Prepared By\nName: {report['full_name']}\nDate: {report['report_date']}",
            "Reviewed By\nName: ____________________\nDate: ____________________",
            "Approved By\nName: ____________________\nDate: ____________________",
        ]
    ]
    canvas.table(["Prepared By", "Reviewed By", "Approved By"], signature_rows, [170, 170, 171])
    canvas.finish()
    write_pdf_document(path, canvas.pages, canvas.logo_meta)


def write_pdf_document(path, pages, logo_meta=None):
    page_width = 595
    page_height = 842

    objects = []
    logo_obj_id = None

    def add_object(content):
        objects.append(content)
        return len(objects)

    font_regular = add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    font_bold = add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")

    if logo_meta:
        w, h, compressed = logo_meta
        logo_obj_id = add_object(
            b"<< /Type /XObject /Subtype /Image /Width "
            + str(w).encode()
            + b" /Height "
            + str(h).encode()
            + b" /ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /FlateDecode /Length "
            + str(len(compressed)).encode()
            + b" >>\nstream\n"
            + compressed
            + b"\nendstream"
        )

    page_ids = []
    for page in pages:
        stream = "\n".join(page).encode("latin-1", "replace")
        resources = f"<< /Font << /F1 {font_regular} 0 R /F2 {font_bold} 0 R >>"
        if logo_obj_id:
            resources += f" /XObject << /Logo {logo_obj_id} 0 R >>"
        resources += " >>"
        content_id = add_object(
            b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream"
        )
        page_id = add_object(
            f"<< /Type /Page /Parent {{PAGES}} 0 R /MediaBox [0 0 {page_width} {page_height}] /Resources {resources} /Contents {content_id} 0 R >>".encode()
        )
        page_ids.append(page_id)

    pages_obj_id = len(objects) + 1
    for page_id in page_ids:
        objects[page_id - 1] = objects[page_id - 1].replace(b"{PAGES}", str(pages_obj_id).encode())
    kids = " ".join(f"{pid} 0 R" for pid in page_ids)
    add_object(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode())
    catalog_id = add_object(f"<< /Type /Catalog /Pages {pages_obj_id} 0 R >>".encode())

    output = [b"%PDF-1.4\n"]
    offsets = [0]
    for idx, obj in enumerate(objects, 1):
        offsets.append(sum(len(part) for part in output))
        output.append(f"{idx} 0 obj\n".encode())
        output.append(obj)
        output.append(b"\nendobj\n")
    xref_offset = sum(len(part) for part in output)
    output.append(f"xref\n0 {len(objects) + 1}\n".encode())
    output.append(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        output.append(f"{off:010d} 00000 n \n".encode())
    output.append(
        f"trailer\n<< /Size {len(objects) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode()
    )
    path.write_bytes(b"".join(output))


@app.route("/")
def index():
    if g.user:
        return redirect(url_for("dashboard"))
    return render_template("index.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    reg = get_db().execute("SELECT value FROM settings WHERE key = 'registration_open'").fetchone()
    if reg and reg["value"] != "1":
        flash("Registration is currently closed.", "warning")
        return redirect(url_for("login"))
    if request.method == "POST":
        csrf_protect()
        full_name = clean_text(request.form.get("full_name"), 120)
        email = clean_text(request.form.get("email"), 160).lower()
        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
            flash("Enter a valid email address.", "danger")
            return redirect(url_for("register"))
        dept = request.form.get("department", "")
        if dept not in DEPARTMENTS:
            flash("Select a valid department.", "danger")
            return redirect(url_for("register"))
        branch = request.form.get("branch", "")
        if branch not in BRANCHES:
            flash("Select a valid branch.", "danger")
            return redirect(url_for("register"))
        password = request.form.get("password", "")
        if len(password) < 8:
            flash("Use a password with at least 8 characters.", "danger")
            return redirect(url_for("register"))
        try:
            get_db().execute(
                """
                INSERT INTO users (
                    full_name, email, phone, department, position, branch,
                    password_hash, role, is_active, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    full_name,
                    email,
                    clean_text(request.form.get("phone"), 40),
                    dept,
                    clean_text(request.form.get("position"), 100),
                    branch,
                    generate_password_hash(password),
                    "employee",
                    0,
                    now(),
                ),
            )
            get_db().commit()
            flash("Account created. Your department admin must activate it before you can sign in.", "info")
            return redirect(url_for("login"))
        except sqlite3.IntegrityError:
            flash("That email address is already registered.", "danger")
    return render_template("register.html", departments=DEPARTMENTS, branches=BRANCHES)


@app.route("/login", methods=["GET", "POST"])
@rate_limit("login", LOGIN_RATE_LIMIT_MAX, LOGIN_RATE_LIMIT_WINDOW)
def login():
    if request.method == "POST":
        csrf_protect()
        email = clean_text(request.form.get("email"), 160).lower()
        password = request.form.get("password", "")
        if check_account_locked(email):
            flash("Account temporarily locked due to too many failed attempts. Try again later.", "danger")
            log.warning("Locked login attempt for %s", email)
            return redirect(url_for("login"))
        user = get_db().execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            locked = row_get(user, "locked_at", None)
            if locked:
                flash("Your account has been locked due to inactivity. Contact your administrator.", "warning")
                return redirect(url_for("login"))
            if not user["is_active"]:
                flash("Your account is pending activation by your department admin.", "warning")
                return redirect(url_for("login"))
            try:
                last_login = user["last_login"]
            except (KeyError, IndexError, TypeError):
                last_login = None
            if last_login:
                try:
                    last_dt = datetime.strptime(last_login, "%Y-%m-%d %H:%M:%S")
                    if (datetime.now() - last_dt) > timedelta(days=INACTIVITY_LOCK_DAYS) and user["role"] not in ("admin", "principal", "superadmin", "shadowadmin"):
                        db = get_db()
                        db.execute(
                            "UPDATE users SET is_active = 0, locked_at = ?, lock_reason = ? WHERE id = ?",
                            (now(), f"Auto-locked: inactive for {INACTIVITY_LOCK_DAYS}+ days", user["id"]),
                        )
                        db.execute(
                            """INSERT INTO account_events (user_id, actor_id, event, reason, created_at)
                               VALUES (?, NULL, 'auto_locked', ?, ?)""",
                            (user["id"], f"Inactive for {INACTIVITY_LOCK_DAYS}+ days", now()),
                        )
                        db.commit()
                        flash(f"Your account has been locked due to {INACTIVITY_LOCK_DAYS} days of inactivity. Contact your administrator.", "warning")
                        return redirect(url_for("login"))
                except (ValueError, TypeError):
                    pass
            clear_login_attempts(email)
            session.clear()
            session["user_id"] = user["id"]
            session["_last_active"] = now()
            get_db().execute("UPDATE users SET last_login = ? WHERE id = ?", (now(), user["id"]))
            get_db().commit()
            session.permanent = True
            flash("Signed in successfully.", "success")
            return redirect(url_for("dashboard"))
        record_failed_attempt(email)
        flash("Invalid email or password.", "danger")
    return render_template("login.html")


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        csrf_protect()
        current = request.form.get("current_password", "")
        new_pass = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        user = get_db().execute("SELECT * FROM users WHERE id = ?", (g.user["id"],)).fetchone()
        if not check_password_hash(user["password_hash"], current):
            flash("Current password is incorrect.", "danger")
            return render_template("change_password.html")
        if len(new_pass) < 8:
            flash("Use a password with at least 8 characters.", "danger")
            return render_template("change_password.html")
        if new_pass != confirm:
            flash("Passwords do not match.", "danger")
            return render_template("change_password.html")
        get_db().execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_pass), g.user["id"]))
        get_db().commit()
        flash("Password changed successfully.", "success")
        return redirect(url_for("dashboard"))
    return render_template("change_password.html")


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    db = get_db()
    if request.method == "POST":
        csrf_protect()
        full_name = clean_text(request.form.get("full_name"), 120)
        phone = clean_text(request.form.get("phone"), 40)
        position = clean_text(request.form.get("position"), 100)
        if not full_name:
            flash("Name is required.", "danger")
            return redirect(url_for("profile"))
        db.execute(
            "UPDATE users SET full_name = ?, phone = ?, position = ? WHERE id = ?",
            (full_name, phone, position, g.user["id"]),
        )
        db.commit()
        g.user = current_user()
        flash("Profile updated.", "success")
        return redirect(url_for("profile"))
    return render_template("profile.html", departments=DEPARTMENTS, branches=BRANCHES)


@app.route("/my-drafts")
@login_required
def my_drafts():
    drafts = get_db().execute(
        "SELECT id, title, report_date, department, updated_at FROM report_drafts WHERE user_id = ? ORDER BY updated_at DESC",
        (g.user["id"],),
    ).fetchall()
    return render_template("my_drafts.html", drafts=drafts, max_drafts=2)


@app.route("/api/reports/drafts", methods=["GET"])
@login_required
def list_drafts():
    drafts = get_db().execute(
        "SELECT id, title, report_date, department, updated_at FROM report_drafts WHERE user_id = ? ORDER BY updated_at DESC",
        (g.user["id"],),
    ).fetchall()
    return [dict(d) for d in drafts]


@app.route("/api/reports/draft", methods=["GET", "POST", "DELETE"])
@login_required
def draft_report():
    db = get_db()
    if request.method == "GET":
        draft_id = request.args.get("id", "")
        draft = db.execute(
            "SELECT data FROM report_drafts WHERE id = ? AND user_id = ?",
            (draft_id, g.user["id"]),
        ).fetchone()
        if draft:
            return json.loads(draft["data"])
        return {"_empty": True}
    is_delete = request.method == "DELETE" or request.form.get("_method") == "DELETE"
    if is_delete:
        csrf_protect()
        draft_id = (request.form.get("id") or "").strip()
        if draft_id:
            db.execute("DELETE FROM report_drafts WHERE id = ? AND user_id = ?", (draft_id, g.user["id"]))
            db.commit()
        if request.method == "POST":
            return redirect(url_for("my_drafts"))
        return {"ok": True}
    if request.method == "POST":
        csrf_protect()
        data = request.get_json(silent=True) or {}
        if not data.get("report_date"):
            return {"ok": False, "error": "report_date required"}, 400
        draft_id = data.get("draft_id")
        count = db.execute("SELECT COUNT(*) AS c FROM report_drafts WHERE user_id = ?", (g.user["id"],)).fetchone()["c"]
        title = data.get("title", "").strip() or f"Draft — {data['report_date']}"
        if draft_id:
            existing = db.execute("SELECT id FROM report_drafts WHERE id = ? AND user_id = ?", (draft_id, g.user["id"])).fetchone()
            if not existing:
                return {"ok": False, "error": "draft not found"}, 404
            db.execute(
                "UPDATE report_drafts SET title = ?, report_date = ?, department = ?, data = ?, updated_at = ? WHERE id = ?",
                (title, data["report_date"], data.get("department", ""), json.dumps(data), now(), draft_id),
            )
        else:
            if count >= 2:
                oldest = db.execute(
                    "SELECT id FROM report_drafts WHERE user_id = ? ORDER BY updated_at ASC LIMIT 1",
                    (g.user["id"],),
                ).fetchone()
                draft_id = oldest["id"]
                db.execute(
                    "UPDATE report_drafts SET title = ?, report_date = ?, department = ?, data = ?, updated_at = ? WHERE id = ?",
                    (title, data["report_date"], data.get("department", ""), json.dumps(data), now(), draft_id),
                )
            else:
                cursor = db.execute(
                    "INSERT INTO report_drafts (user_id, title, report_date, department, data, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (g.user["id"], title, data["report_date"], data.get("department", ""), json.dumps(data), now()),
                )
                draft_id = cursor.lastrowid
        db.commit()
        return {"ok": True, "draft_id": draft_id, "title": title}


@app.route("/logout")
def logout():
    session.clear()
    flash("Signed out.", "info")
    return redirect(url_for("index"))


@app.route("/dashboard")
@app.route("/dashboard/<int:archived>")
@login_required
def dashboard(archived=0):
    db = get_db()
    show_archived = 1 if archived == 1 else 0
    filters = report_filter_values()
    reports = apply_report_filters(visible_report_rows(db, show_archived), filters)
    # Stats cover every role-scoped report (no truncation), so "Total Reports"
    # and every derived rate are accurate; the rendered list is capped for size.
    insights = build_dashboard_insights(reports, get_visible_staff_count(db))
    modules = expansion_modules(insights)
    return render_template(
        "dashboard.html",
        reports=reports[:DASHBOARD_LIST_CAP],
        show_archived=bool(show_archived),
        insights=insights,
        modules=modules,
        filters=filters,
        filter_users=scoped_users(),
        departments=DEPARTMENTS,
    )


@app.route("/modules")
@login_required
def modules():
    db = get_db()
    archived_expr = archived_filter("reports.archived")
    if g.user["role"] == "employee":
        reports = db.execute(
            f"SELECT * FROM reports WHERE user_id = ? AND {archived_filter()} = 0 ORDER BY report_date DESC, id DESC",
            (g.user["id"],),
        ).fetchall()
    elif g.user["role"] == "admin":
        reports = db.execute(
            f"""
            SELECT reports.*, users.full_name
            FROM reports
            JOIN users ON users.id = reports.user_id
            JOIN report_access ON report_access.employee_id = reports.user_id
            WHERE report_access.admin_id = ? AND {archived_expr} = 0
            ORDER BY reports.report_date DESC, reports.id DESC
            """,
            (g.user["id"],),
        ).fetchall()
    else:
        reports = db.execute(
            f"""
            SELECT reports.*, users.full_name
            FROM reports
            JOIN users ON users.id = reports.user_id
            WHERE {archived_expr} = 0
            ORDER BY reports.report_date DESC, reports.id DESC
            """
        ).fetchall()
    insights = build_dashboard_insights(reports, get_visible_staff_count(db))
    return render_template("modules.html", modules=expansion_modules(insights), insights=insights)


ADMISSION_STATUSES = ("pending", "needs_info", "verified", "rejected")


def can_view_admission(user, admission):
    """Keep admission records within the creator's reporting scope."""
    if not admission:
        return False
    if user["role"] in EXECUTIVE_ROLES:
        return True
    if admission["created_by"] == user["id"]:
        return True
    if user["role"] == "admin":
        allowed = get_db().execute(
            "SELECT 1 FROM report_access WHERE admin_id = ? AND employee_id = ?",
            (user["id"], admission["created_by"]),
        ).fetchone()
        return bool(allowed)
    return False


def get_admission_or_404(admission_id):
    admission = get_db().execute(
        """SELECT admissions.*, users.full_name AS creator_name
           FROM admissions JOIN users ON users.id = admissions.created_by
           WHERE admissions.id = ?""",
        (admission_id,),
    ).fetchone()
    if not admission or not can_view_admission(g.user, admission):
        abort(404)
    return admission


@app.route("/admissions", methods=["GET", "POST"])
@login_required
def admissions():
    db = get_db()
    if request.method == "POST":
        csrf_protect()
        applicant_name = clean_text(request.form.get("applicant_name"), 160)
        course = clean_text(request.form.get("course"), 160)
        admission_date = clean_text(request.form.get("admission_date"), 10)
        if not applicant_name or not course:
            flash("Applicant name and course are required.", "danger")
            return redirect(url_for("admissions"))
        if admission_date and not parse_report_date(admission_date):
            flash("Use a valid admission date.", "danger")
            return redirect(url_for("admissions"))
        timestamp = now()
        cursor = db.execute(
            """INSERT INTO admissions
               (applicant_name, contact, course, intake, admission_date, source, notes, status, created_by, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?)""",
            (
                applicant_name,
                clean_text(request.form.get("contact"), 80),
                course,
                clean_text(request.form.get("intake"), 80),
                admission_date,
                clean_text(request.form.get("source"), 120),
                clean_text(request.form.get("notes"), 2000),
                g.user["id"],
                timestamp,
                timestamp,
            ),
        )
        admission_id = cursor.lastrowid
        db.execute(
            """INSERT INTO admission_verification_history
               (admission_id, reviewer_id, status, notes, created_at)
               VALUES (?, ?, 'pending', ?, ?)""",
            (admission_id, g.user["id"], "Record captured; awaiting verification.", timestamp),
        )
        db.commit()
        flash("Admission record captured and queued for verification.", "success")
        return redirect(url_for("admission_detail", admission_id=admission_id))

    if g.user["role"] in EXECUTIVE_ROLES:
        records = db.execute(
            """SELECT admissions.*, users.full_name AS creator_name
               FROM admissions JOIN users ON users.id = admissions.created_by
               ORDER BY admissions.created_at DESC, admissions.id DESC"""
        ).fetchall()
    elif g.user["role"] == "admin":
        records = db.execute(
            """SELECT DISTINCT admissions.*, users.full_name AS creator_name
               FROM admissions JOIN users ON users.id = admissions.created_by
               LEFT JOIN report_access ON report_access.employee_id = admissions.created_by
               WHERE admissions.created_by = ? OR report_access.admin_id = ?
               ORDER BY admissions.created_at DESC, admissions.id DESC""",
            (g.user["id"], g.user["id"]),
        ).fetchall()
    else:
        records = db.execute(
            """SELECT admissions.*, users.full_name AS creator_name
               FROM admissions JOIN users ON users.id = admissions.created_by
               WHERE admissions.created_by = ?
               ORDER BY admissions.created_at DESC, admissions.id DESC""",
            (g.user["id"],),
        ).fetchall()
    counts = {status: 0 for status in ADMISSION_STATUSES}
    for record in records:
        counts[record["status"]] = counts.get(record["status"], 0) + 1
    return render_template("admissions.html", admissions=records, counts=counts, statuses=ADMISSION_STATUSES)


@app.route("/admissions/<int:admission_id>")
@login_required
def admission_detail(admission_id):
    admission = get_admission_or_404(admission_id)
    history = get_db().execute(
        """SELECT h.*, u.full_name AS reviewer_name
           FROM admission_verification_history h
           LEFT JOIN users u ON u.id = h.reviewer_id
           WHERE h.admission_id = ?
           ORDER BY h.created_at ASC, h.id ASC""",
        (admission_id,),
    ).fetchall()
    return render_template("admission_detail.html", admission=admission, history=history, statuses=ADMISSION_STATUSES)


@app.route("/admissions/<int:admission_id>/verify", methods=["POST"])
@admin_required
def verify_admission(admission_id):
    csrf_protect()
    admission = get_admission_or_404(admission_id)
    status = clean_text(request.form.get("status"), 30)
    notes = clean_text(request.form.get("notes"), 2000)
    if status not in ADMISSION_STATUSES or status == "pending":
        abort(400)
    if status in ("needs_info", "rejected") and not notes:
        flash("Add a note explaining the verification decision.", "danger")
        return redirect(url_for("admission_detail", admission_id=admission_id))
    timestamp = now()
    db = get_db()
    db.execute(
        "UPDATE admissions SET status = ?, notes = ?, updated_at = ? WHERE id = ?",
        (status, notes, timestamp, admission_id),
    )
    db.execute(
        """INSERT INTO admission_verification_history
           (admission_id, reviewer_id, status, notes, created_at)
           VALUES (?, ?, ?, ?, ?)""",
        (admission_id, g.user["id"], status, notes, timestamp),
    )
    db.commit()
    if admission["created_by"] != g.user["id"]:
        create_notification(
            user_id=admission["created_by"],
            kind="admission",
            title=f"Admission marked {status.replace('_', ' ')}",
            body=f"The admission record for {admission['applicant_name']} was updated by {g.user['full_name']}.",
            link=url_for("admission_detail", admission_id=admission_id),
            actor_id=g.user["id"],
        )
    flash("Admission verification status updated.", "success")
    return redirect(url_for("admission_detail", admission_id=admission_id))


@app.route("/admissions/<int:admission_id>/payment", methods=["POST"])
@admin_required
def update_admission_payment(admission_id):
    csrf_protect()
    admission = get_admission_or_404(admission_id)
    action = clean_text(request.form.get("action"), 20)
    note = clean_text(request.form.get("note"), 1000)
    timestamp = now()
    db = get_db()
    if action == "mark_paid":
        reference = clean_text(request.form.get("payment_reference"), 120)
        if not reference:
            flash("Record the approved payment reference before marking fees paid.", "danger")
            return redirect(url_for("admission_detail", admission_id=admission_id))
        db.execute(
            """UPDATE admissions SET fee_status = 'paid', fee_paid_at = ?,
                   fee_payment_reference = ?, updated_at = ? WHERE id = ?""",
            (timestamp, reference, timestamp, admission_id),
        )
        history_status = "fee_paid"
        history_note = note or f"Fee payment recorded under reference {reference}."
    elif action == "reopen":
        if not note:
            flash("Explain why the fee-paid state is being reopened.", "danger")
            return redirect(url_for("admission_detail", admission_id=admission_id))
        db.execute(
            """UPDATE admissions SET fee_status = 'unpaid', fee_paid_at = NULL,
                   fee_payment_reference = '', updated_at = ? WHERE id = ?""",
            (timestamp, admission_id),
        )
        history_status = "fee_reopened"
        history_note = note
    else:
        abort(400)
    db.execute(
        """INSERT INTO admission_verification_history
           (admission_id, reviewer_id, status, notes, created_at)
           VALUES (?, ?, ?, ?, ?)""",
        (admission_id, g.user["id"], history_status, history_note, timestamp),
    )
    db.commit()
    if admission["created_by"] != g.user["id"]:
        create_notification(
            user_id=admission["created_by"],
            kind="admission",
            title="Admission fee state updated",
            body=f"The fee state for {admission['applicant_name']} was updated.",
            link=url_for("admission_detail", admission_id=admission_id),
            actor_id=g.user["id"],
        )
    flash("Admission fee state updated.", "success")
    return redirect(url_for("admission_detail", admission_id=admission_id))


def scoped_users(include_self=True):
    """Return active users the signed-in user is allowed to coordinate."""
    db = get_db()
    if g.user["role"] in EXECUTIVE_ROLES:
        rows = db.execute(
            """SELECT id, full_name, department, position, role
               FROM users WHERE is_active = 1 AND deleted_at IS NULL AND """ + hide_shadow_sql() +
            " ORDER BY department, full_name"
        ).fetchall()
    elif g.user["role"] == "admin":
        rows = db.execute(
            """SELECT DISTINCT users.id, users.full_name, users.department, users.position, users.role
               FROM users LEFT JOIN report_access ON report_access.employee_id = users.id
               WHERE users.is_active = 1 AND users.deleted_at IS NULL
                 AND (users.id = ? OR report_access.admin_id = ?)
               ORDER BY users.department, users.full_name""",
            (g.user["id"], g.user["id"]),
        ).fetchall()
    else:
        rows = db.execute(
            """SELECT id, full_name, department, position, role
               FROM users WHERE id = ? AND is_active = 1 AND deleted_at IS NULL""",
            (g.user["id"],),
        ).fetchall()
    if include_self:
        return rows
    return [row for row in rows if row["id"] != g.user["id"]]


def can_coordinate_user(user_id):
    return any(row["id"] == user_id for row in scoped_users())


@app.route("/intake-targets", methods=["GET", "POST"])
@login_required
def intake_targets():
    db = get_db()
    officers = scoped_users()
    if request.method == "POST":
        csrf_protect()
        if g.user["role"] not in ADMIN_ROLES:
            abort(403)
        try:
            officer_id = int(request.form.get("officer_id", ""))
            target_count = int(request.form.get("target_count", ""))
        except (TypeError, ValueError):
            flash("Choose an officer and enter a valid target.", "danger")
            return redirect(url_for("intake_targets"))
        target_month = clean_text(request.form.get("target_month"), 7)
        try:
            datetime.strptime(target_month, "%Y-%m")
        except (TypeError, ValueError):
            flash("Choose a valid target month.", "danger")
            return redirect(url_for("intake_targets"))
        if target_count < 1 or target_count > 10000:
            flash("Target must be between 1 and 10,000 admissions.", "danger")
            return redirect(url_for("intake_targets"))
        if not can_coordinate_user(officer_id):
            abort(403)
        timestamp = now()
        db.execute(
            """INSERT INTO intake_targets
               (officer_id, target_month, target_count, created_by, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?)
               ON CONFLICT(officer_id, target_month) DO UPDATE SET
                   target_count = excluded.target_count,
                   created_by = excluded.created_by,
                   updated_at = excluded.updated_at""",
            (officer_id, target_month, target_count, g.user["id"], timestamp, timestamp),
        )
        db.commit()
        if officer_id != g.user["id"]:
            create_notification(
                user_id=officer_id,
                kind="intake_target",
                title=f"Intake target set for {target_month}",
                body=f"Your admissions target is {target_count} for {target_month}.",
                link=url_for("intake_targets"),
                actor_id=g.user["id"],
            )
        flash("Monthly intake target saved.", "success")
        return redirect(url_for("intake_targets", month=target_month))

    target_month = clean_text(request.args.get("month"), 7) or datetime.now().strftime("%Y-%m")
    try:
        datetime.strptime(target_month, "%Y-%m")
    except (TypeError, ValueError):
        target_month = datetime.now().strftime("%Y-%m")
    rows = []
    for officer in officers:
        target = db.execute(
            "SELECT * FROM intake_targets WHERE officer_id = ? AND target_month = ?",
            (officer["id"], target_month),
        ).fetchone()
        actual_row = db.execute(
            """SELECT COUNT(*) AS c, MAX(updated_at) AS last_activity
               FROM admissions
               WHERE created_by = ? AND status = 'verified' AND fee_status = 'paid'
                 AND substr(COALESCE(NULLIF(admission_date, ''), created_at), 1, 7) = ?""",
            (officer["id"], target_month),
        ).fetchone()
        target_count = target["target_count"] if target else 0
        actual = actual_row["c"] if actual_row else 0
        rows.append({
            "officer": officer,
            "target": target_count,
            "actual": actual,
            "remaining": max(target_count - actual, 0),
            "percentage": min(int((actual / target_count) * 100), 100) if target_count else 0,
                "last_activity": actual_row["last_activity"] if actual_row else None,
        })
    rows.sort(key=lambda item: (item["percentage"], item["actual"]), reverse=True)
    return render_template(
        "intake_targets.html",
        target_month=target_month,
        officers=officers,
        rows=rows,
        total_target=sum(item["target"] for item in rows),
        total_actual=sum(item["actual"] for item in rows),
    )


MEETING_ACTION_STATUSES = ("open", "in_progress", "blocked", "completed")


def can_view_meeting(user, meeting):
    if not meeting:
        return False
    if user["role"] in EXECUTIVE_ROLES:
        return True
    if meeting["created_by"] == user["id"]:
        return True
    if user["role"] == "admin":
        creator_allowed = get_db().execute(
            "SELECT 1 FROM report_access WHERE admin_id = ? AND employee_id = ?",
            (user["id"], meeting["created_by"]),
        ).fetchone()
        if creator_allowed:
            return True
    assigned = get_db().execute(
        "SELECT 1 FROM meeting_actions WHERE meeting_id = ? AND owner_id = ?",
        (meeting["id"], user["id"]),
    ).fetchone()
    return bool(assigned)


def get_meeting_or_404(meeting_id):
    meeting = get_db().execute(
        """SELECT meetings.*, users.full_name AS creator_name
           FROM meetings JOIN users ON users.id = meetings.created_by
           WHERE meetings.id = ?""",
        (meeting_id,),
    ).fetchone()
    if not meeting or not can_view_meeting(g.user, meeting):
        abort(404)
    return meeting


@app.route("/meetings", methods=["GET", "POST"])
@login_required
def meetings():
    db = get_db()
    if request.method == "POST":
        csrf_protect()
        title = clean_text(request.form.get("title"), 180)
        meeting_date = clean_text(request.form.get("meeting_date"), 10)
        summary = clean_text(request.form.get("summary"), 4000)
        if not title or not parse_report_date(meeting_date) or not summary:
            flash("Meeting title, valid date and summary are required.", "danger")
            return redirect(url_for("meetings"))
        timestamp = now()
        cursor = db.execute(
            """INSERT INTO meetings
               (title, meeting_date, department, attendees, summary, created_by, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                title,
                meeting_date,
                clean_text(request.form.get("department"), 100) or g.user["department"],
                clean_text(request.form.get("attendees"), 2000),
                summary,
                g.user["id"],
                timestamp,
                timestamp,
            ),
        )
        meeting_id = cursor.lastrowid
        db.commit()
        flash("Meeting record saved. Add accountable actions next.", "success")
        return redirect(url_for("meeting_detail", meeting_id=meeting_id))

    if g.user["role"] in EXECUTIVE_ROLES:
        records = db.execute(
            """SELECT meetings.*, users.full_name AS creator_name,
                      COUNT(meeting_actions.id) AS action_count,
                      SUM(CASE WHEN meeting_actions.status != 'completed' THEN 1 ELSE 0 END) AS open_count
               FROM meetings JOIN users ON users.id = meetings.created_by
               LEFT JOIN meeting_actions ON meeting_actions.meeting_id = meetings.id
               GROUP BY meetings.id, users.full_name
               ORDER BY meetings.meeting_date DESC, meetings.id DESC"""
        ).fetchall()
    elif g.user["role"] == "admin":
        records = db.execute(
            """SELECT meetings.*, users.full_name AS creator_name,
                      COUNT(DISTINCT meeting_actions.id) AS action_count,
                      COUNT(DISTINCT CASE WHEN meeting_actions.status != 'completed' THEN meeting_actions.id END) AS open_count
               FROM meetings JOIN users ON users.id = meetings.created_by
               LEFT JOIN report_access ON report_access.employee_id = meetings.created_by
               LEFT JOIN meeting_actions ON meeting_actions.meeting_id = meetings.id
               WHERE meetings.created_by = ? OR report_access.admin_id = ? OR meeting_actions.owner_id = ?
               GROUP BY meetings.id, users.full_name
               ORDER BY meetings.meeting_date DESC, meetings.id DESC""",
            (g.user["id"], g.user["id"], g.user["id"]),
        ).fetchall()
    else:
        records = db.execute(
            """SELECT meetings.*, users.full_name AS creator_name,
                      COUNT(DISTINCT meeting_actions.id) AS action_count,
                      COUNT(DISTINCT CASE WHEN meeting_actions.status != 'completed' THEN meeting_actions.id END) AS open_count
               FROM meetings JOIN users ON users.id = meetings.created_by
               LEFT JOIN meeting_actions ON meeting_actions.meeting_id = meetings.id
               WHERE meetings.created_by = ? OR meeting_actions.owner_id = ?
               GROUP BY meetings.id, users.full_name
               ORDER BY meetings.meeting_date DESC, meetings.id DESC""",
            (g.user["id"], g.user["id"]),
        ).fetchall()
    return render_template("meetings.html", meetings=records, departments=DEPARTMENTS)


@app.route("/meetings/<int:meeting_id>")
@login_required
def meeting_detail(meeting_id):
    meeting = get_meeting_or_404(meeting_id)
    actions = get_db().execute(
        """SELECT meeting_actions.*, users.full_name AS owner_name,
                  creator.full_name AS action_creator_name
           FROM meeting_actions
           JOIN users ON users.id = meeting_actions.owner_id
           JOIN users creator ON creator.id = meeting_actions.created_by
           WHERE meeting_actions.meeting_id = ?
           ORDER BY meeting_actions.due_date ASC, meeting_actions.id ASC""",
        (meeting_id,),
    ).fetchall()
    can_add_actions = g.user["role"] in ADMIN_ROLES or meeting["created_by"] == g.user["id"]
    return render_template(
        "meeting_detail.html",
        meeting=meeting,
        actions=actions,
        owners=scoped_users(),
        action_statuses=MEETING_ACTION_STATUSES,
        can_add_actions=can_add_actions,
    )


@app.route("/meetings/<int:meeting_id>/actions", methods=["POST"])
@login_required
def add_meeting_action(meeting_id):
    csrf_protect()
    meeting = get_meeting_or_404(meeting_id)
    if g.user["role"] not in ADMIN_ROLES and meeting["created_by"] != g.user["id"]:
        abort(403)
    description = clean_text(request.form.get("description"), 1000)
    due_date = clean_text(request.form.get("due_date"), 10)
    try:
        owner_id = int(request.form.get("owner_id", ""))
    except (TypeError, ValueError):
        owner_id = 0
    if not description or not owner_id or not can_coordinate_user(owner_id):
        flash("Action description and an authorised owner are required.", "danger")
        return redirect(url_for("meeting_detail", meeting_id=meeting_id))
    if due_date and not parse_report_date(due_date):
        flash("Use a valid due date.", "danger")
        return redirect(url_for("meeting_detail", meeting_id=meeting_id))
    timestamp = now()
    get_db().execute(
        """INSERT INTO meeting_actions
           (meeting_id, description, owner_id, due_date, status, created_by, created_at, updated_at)
           VALUES (?, ?, ?, ?, 'open', ?, ?, ?)""",
        (meeting_id, description, owner_id, due_date, g.user["id"], timestamp, timestamp),
    )
    get_db().commit()
    if owner_id != g.user["id"]:
        create_notification(
            user_id=owner_id,
            kind="meeting_action",
            title=f"New action: {meeting['title']}",
            body=description,
            link=url_for("meeting_detail", meeting_id=meeting_id),
            actor_id=g.user["id"],
        )
    flash("Meeting action assigned.", "success")
    return redirect(url_for("meeting_detail", meeting_id=meeting_id))


@app.route("/meetings/<int:meeting_id>/actions/<int:action_id>/status", methods=["POST"])
@login_required
def update_meeting_action(meeting_id, action_id):
    csrf_protect()
    meeting = get_meeting_or_404(meeting_id)
    action = get_db().execute(
        "SELECT * FROM meeting_actions WHERE id = ? AND meeting_id = ?",
        (action_id, meeting_id),
    ).fetchone()
    if not action:
        abort(404)
    if action["owner_id"] != g.user["id"] and g.user["role"] not in ADMIN_ROLES:
        abort(403)
    status = clean_text(request.form.get("status"), 30)
    notes = clean_text(request.form.get("completion_notes"), 2000)
    if status not in MEETING_ACTION_STATUSES:
        abort(400)
    if status == "blocked" and not notes:
        flash("Explain what is blocking this action.", "danger")
        return redirect(url_for("meeting_detail", meeting_id=meeting_id))
    timestamp = now()
    completed_at = timestamp if status == "completed" else None
    get_db().execute(
        """UPDATE meeting_actions
           SET status = ?, completion_notes = ?, completed_at = ?, updated_at = ?
           WHERE id = ?""",
        (status, notes, completed_at, timestamp, action_id),
    )
    get_db().commit()
    if meeting["created_by"] != g.user["id"]:
        create_notification(
            user_id=meeting["created_by"],
            kind="meeting_action",
            title=f"Action marked {status.replace('_', ' ')}",
            body=action["description"],
            link=url_for("meeting_detail", meeting_id=meeting_id),
            actor_id=g.user["id"],
        )
    flash("Action status updated.", "success")
    return redirect(url_for("meeting_detail", meeting_id=meeting_id))


INCENTIVE_STATUSES = ("proposed", "approved", "rejected", "paid")


def can_view_incentive(user, incentive):
    if not incentive:
        return False
    if user["role"] in EXECUTIVE_ROLES:
        return True
    if incentive["employee_id"] == user["id"]:
        return True
    if user["role"] == "admin":
        return bool(get_db().execute(
            "SELECT 1 FROM report_access WHERE admin_id = ? AND employee_id = ?",
            (user["id"], incentive["employee_id"]),
        ).fetchone())
    return False


def visible_incentives():
    db = get_db()
    if g.user["role"] in EXECUTIVE_ROLES:
        return db.execute(
            """SELECT incentives.*, users.full_name AS employee_name, users.department
               FROM incentives JOIN users ON users.id = incentives.employee_id
               WHERE """ + hide_shadow_sql() +
            " ORDER BY incentives.period_month DESC, incentives.id DESC"
        ).fetchall()
    if g.user["role"] == "admin":
        return db.execute(
            """SELECT DISTINCT incentives.*, users.full_name AS employee_name, users.department
               FROM incentives JOIN users ON users.id = incentives.employee_id
               LEFT JOIN report_access ON report_access.employee_id = incentives.employee_id
               WHERE incentives.employee_id = ? OR report_access.admin_id = ?
               ORDER BY incentives.period_month DESC, incentives.id DESC""",
            (g.user["id"], g.user["id"]),
        ).fetchall()
    return db.execute(
        """SELECT incentives.*, users.full_name AS employee_name, users.department
           FROM incentives JOIN users ON users.id = incentives.employee_id
           WHERE incentives.employee_id = ?
           ORDER BY incentives.period_month DESC, incentives.id DESC""",
        (g.user["id"],),
    ).fetchall()


@app.route("/incentives", methods=["GET", "POST"])
@login_required
def incentives():
    db = get_db()
    if request.method == "POST":
        csrf_protect()
        if g.user["role"] not in ("admin", "superadmin", "shadowadmin"):
            abort(403)
        try:
            employee_id = int(request.form.get("employee_id", ""))
            units = int(request.form.get("units", ""))
            rate = Decimal(clean_text(request.form.get("rate"), 30))
        except (TypeError, ValueError, InvalidOperation):
            flash("Choose an employee and enter valid units and rate.", "danger")
            return redirect(url_for("incentives"))
        period_month = clean_text(request.form.get("period_month"), 7)
        description = clean_text(request.form.get("description"), 500)
        try:
            datetime.strptime(period_month, "%Y-%m")
        except (TypeError, ValueError):
            flash("Choose a valid incentive month.", "danger")
            return redirect(url_for("incentives"))
        if not can_coordinate_user(employee_id):
            abort(403)
        if not description or units < 1 or units > 100000 or rate <= 0 or rate > Decimal("10000000"):
            flash("Description, positive units and a reasonable rate are required.", "danger")
            return redirect(url_for("incentives"))
        rate_cents = int((rate * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        amount_cents = units * rate_cents
        timestamp = now()
        cursor = db.execute(
            """INSERT INTO incentives
               (employee_id, period_month, description, units, rate_cents, amount_cents,
                status, notes, created_by, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, 'proposed', ?, ?, ?, ?)""",
            (
                employee_id,
                period_month,
                description,
                units,
                rate_cents,
                amount_cents,
                clean_text(request.form.get("notes"), 2000),
                g.user["id"],
                timestamp,
                timestamp,
            ),
        )
        incentive_id = cursor.lastrowid
        db.execute(
            """INSERT INTO incentive_events (incentive_id, actor_id, event, notes, created_at)
               VALUES (?, ?, 'proposed', ?, ?)""",
            (incentive_id, g.user["id"], description, timestamp),
        )
        db.commit()
        if employee_id != g.user["id"]:
            create_notification(
                user_id=employee_id,
                kind="incentive",
                title=f"Incentive proposed for {period_month}",
                body=f"KES {amount_cents / 100:,.2f} is awaiting independent approval.",
                link=url_for("incentives"),
                actor_id=g.user["id"],
            )
        flash("Incentive proposal saved for independent approval.", "success")
        return redirect(url_for("incentives"))

    records = visible_incentives()
    totals = {status: sum(row["amount_cents"] for row in records if row["status"] == status) for status in INCENTIVE_STATUSES}
    return render_template(
        "incentives.html",
        incentives=records,
        totals=totals,
        employees=scoped_users(),
        current_month=datetime.now().strftime("%Y-%m"),
    )


@app.route("/incentives/<int:incentive_id>/status", methods=["POST"])
@login_required
def update_incentive_status(incentive_id):
    csrf_protect()
    db = get_db()
    incentive = db.execute("SELECT * FROM incentives WHERE id = ?", (incentive_id,)).fetchone()
    if not incentive or not can_view_incentive(g.user, incentive):
        abort(404)
    action = clean_text(request.form.get("action"), 30)
    notes = clean_text(request.form.get("notes"), 2000)
    timestamp = now()
    if action in ("approve", "reject"):
        if g.user["role"] not in EXECUTIVE_ROLES:
            abort(403)
        if incentive["status"] != "proposed":
            abort(400)
        if incentive["created_by"] == g.user["id"]:
            flash("A different authorised executive must review this proposal.", "danger")
            return redirect(url_for("incentives"))
        if action == "reject" and not notes:
            flash("Explain why the incentive is rejected.", "danger")
            return redirect(url_for("incentives"))
        new_status = "approved" if action == "approve" else "rejected"
        db.execute(
            """UPDATE incentives SET status = ?, notes = ?, approved_by = ?, approved_at = ?, updated_at = ?
               WHERE id = ?""",
            (new_status, notes, g.user["id"], timestamp, timestamp, incentive_id),
        )
    elif action == "mark_paid":
        if g.user["role"] not in SUPER_ROLES or incentive["status"] != "approved":
            abort(403)
        payment_reference = clean_text(request.form.get("payment_reference"), 120)
        if not payment_reference:
            flash("A payment reference is required before marking an incentive paid.", "danger")
            return redirect(url_for("incentives"))
        new_status = "paid"
        notes = notes or "Payment recorded."
        db.execute(
            """UPDATE incentives SET status = 'paid', payment_reference = ?, paid_by = ?, paid_at = ?,
                   notes = ?, updated_at = ? WHERE id = ?""",
            (payment_reference, g.user["id"], timestamp, notes, timestamp, incentive_id),
        )
    else:
        abort(400)
    db.execute(
        """INSERT INTO incentive_events (incentive_id, actor_id, event, notes, created_at)
           VALUES (?, ?, ?, ?, ?)""",
        (incentive_id, g.user["id"], new_status, notes, timestamp),
    )
    db.commit()
    if incentive["employee_id"] != g.user["id"]:
        create_notification(
            user_id=incentive["employee_id"],
            kind="incentive",
            title=f"Incentive {new_status}",
            body=f"Your {incentive['period_month']} incentive is now {new_status}.",
            link=url_for("incentives"),
            actor_id=g.user["id"],
        )
    flash(f"Incentive marked {new_status}.", "success")
    return redirect(url_for("incentives"))

OKR_STATUSES = ("draft", "active", "at_risk", "completed", "archived")
OKR_KEY_RESULT_STATUSES = ("draft", "active", "at_risk", "achieved", "closed")
WINGU_DISPATCH_STATUSES = ("ready", "dispatching", "accepted", "rejected", "needs_attention", "cancelled")


@app.route("/okrs", methods=["GET", "POST"])
@login_required
def okrs():
    db = get_db()
    if request.method == "POST":
        csrf_protect()
        if not can_manage_okr(g.user):
            abort(403)
        title = clean_text(request.form.get("title"), 180)
        description = clean_text(request.form.get("description"), 3000)
        department = clean_text(request.form.get("department"), 100)
        period_start = clean_text(request.form.get("period_start"), 10)
        period_end = clean_text(request.form.get("period_end"), 10)
        owner_id = int_field("owner_id") or None
        status = clean_text(request.form.get("status"), 20)
        if g.user["role"] == "admin":
            department = g.user["department"]
        if not title or department not in DEPARTMENTS or status != "draft":
            flash("Provide a title, valid department and valid status.", "danger")
            return redirect(url_for("okrs"))
        start = parse_report_date(period_start)
        end = parse_report_date(period_end)
        if not start or not end or end < start:
            flash("Provide a valid OKR period; the end date must follow the start date.", "danger")
            return redirect(url_for("okrs"))
        owner = None
        if owner_id:
            owner = db.execute("SELECT * FROM users WHERE id = ? AND deleted_at IS NULL", (owner_id,)).fetchone()
            if not owner or (g.user["role"] == "admin" and owner["department"] != g.user["department"]):
                abort(400)
        timestamp = now()
        cursor = db.execute(
            """INSERT INTO okr_objectives
               (title, description, department, period_start, period_end, owner_id, status, created_by, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (title, description, department, period_start, period_end, owner_id, status, g.user["id"], timestamp, timestamp),
        )
        db.commit()
        flash("Objective created. Add measurable key results next.", "success")
        return redirect(url_for("okr_detail", objective_id=cursor.lastrowid))

    objectives = db.execute(
        """SELECT o.*, owner.full_name AS owner_name,
                  COUNT(kr.id) AS key_result_count
           FROM okr_objectives o
           LEFT JOIN users owner ON owner.id = o.owner_id
           LEFT JOIN okr_key_results kr ON kr.objective_id = o.id
           GROUP BY o.id, owner.full_name
           ORDER BY o.period_end ASC, o.id DESC"""
    ).fetchall()
    objectives = [row for row in objectives if can_view_okr(g.user, row)]
    users = strip_hidden_users(db.execute(
        "SELECT id, full_name, department, role FROM users WHERE is_active = 1 AND deleted_at IS NULL ORDER BY full_name"
    ).fetchall())
    if g.user["role"] == "admin":
        users = [user for user in users if user["department"] == g.user["department"]]
    return render_template(
        "okrs.html", objectives=objectives, users=users, departments=DEPARTMENTS,
        can_create=can_manage_okr(g.user), statuses=OKR_STATUSES,
    )


@app.route("/okrs/load-toolkit-draft", methods=["POST"])
@principal_required
def load_toolkit_okr_draft():
    csrf_protect()
    template_path = BASE_DIR / "data" / "toolkit_okr_template.json"
    try:
        template = json.loads(template_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        log.exception("Unable to load Toolkit OKR template")
        flash("The finalized Toolkit OKR template could not be loaded.", "danger")
        return redirect(url_for("okrs"))
    db = get_db()
    timestamp = now()
    objective_count = 0
    result_count = 0
    for objective_data in template.get("objectives", []):
        department = objective_data.get("department", "")
        if department not in DEPARTMENTS:
            log.error("Invalid department in OKR template: %s", department)
            abort(500)
        objective_key = f"{template['template']}:{objective_data['key']}"
        objective = db.execute(
            "SELECT id FROM okr_objectives WHERE template_key = ?", (objective_key,)
        ).fetchone()
        if objective:
            objective_id = objective["id"]
        else:
            objective_id = db.execute(
                """INSERT INTO okr_objectives
                   (title, description, department, period_start, period_end, owner_id, status,
                    created_by, created_at, updated_at, template_key)
                   VALUES (?, ?, ?, ?, ?, NULL, 'draft', ?, ?, ?, ?)""",
                (
                    objective_data["title"], objective_data.get("description", ""), department,
                    template["period_start"], template["period_end"], g.user["id"], timestamp,
                    timestamp, objective_key,
                ),
            ).lastrowid
            objective_count += 1
        for kr_key, title, baseline, target, unit in objective_data.get("key_results", []):
            result_key = f"{objective_key}:{kr_key}"
            if db.execute("SELECT id FROM okr_key_results WHERE template_key = ?", (result_key,)).fetchone():
                continue
            db.execute(
                """INSERT INTO okr_key_results
                   (objective_id, title, baseline, target, current_value, unit, due_date, owner_id,
                    status, created_at, updated_at, template_key)
                   VALUES (?, ?, ?, ?, ?, ?, ?, NULL, 'draft', ?, ?, ?)""",
                (
                    objective_id, title, baseline, target, baseline, unit, template["period_end"],
                    timestamp, timestamp, result_key,
                ),
            )
            result_count += 1
    db.commit()
    if objective_count or result_count:
        flash(f"Loaded {objective_count} finalized portfolio objectives and {result_count} key results for owner assignment.", "success")
    else:
        flash("The finalized Toolkit portfolio is already loaded; no duplicates were created.", "info")
    return redirect(url_for("okrs"))


def get_okr_or_404(objective_id):
    objective = get_db().execute(
        """SELECT o.*, owner.full_name AS owner_name, creator.full_name AS creator_name
           FROM okr_objectives o
           LEFT JOIN users owner ON owner.id = o.owner_id
           JOIN users creator ON creator.id = o.created_by
           WHERE o.id = ?""",
        (objective_id,),
    ).fetchone()
    if not objective:
        abort(404)
    if not can_view_okr(g.user, objective):
        abort(403)
    return objective


@app.route("/okrs/<int:objective_id>")
@login_required
def okr_detail(objective_id):
    db = get_db()
    objective = get_okr_or_404(objective_id)
    key_results = db.execute(
        """SELECT kr.*, owner.full_name AS owner_name
           FROM okr_key_results kr LEFT JOIN users owner ON owner.id = kr.owner_id
           WHERE kr.objective_id = ? ORDER BY kr.due_date, kr.id""",
        (objective_id,),
    ).fetchall()
    rendered_results = []
    for result in key_results:
        updates = db.execute(
            """SELECT u.*, users.full_name FROM okr_updates u JOIN users ON users.id = u.user_id
               WHERE u.key_result_id = ? ORDER BY u.created_at DESC, u.id DESC""",
            (result["id"],),
        ).fetchall()
        rendered_results.append({"record": result, "progress": okr_progress_percent(result), "updates": updates})
    users = strip_hidden_users(db.execute(
        "SELECT id, full_name, department, role FROM users WHERE is_active = 1 AND deleted_at IS NULL ORDER BY full_name"
    ).fetchall())
    if g.user["role"] == "admin":
        users = [user for user in users if user["department"] == g.user["department"]]
    return render_template(
        "okr_detail.html", objective=objective, key_results=rendered_results, users=users,
        can_manage=can_manage_okr(g.user, objective), result_statuses=OKR_KEY_RESULT_STATUSES,
    )


@app.route("/okrs/<int:objective_id>/key-results", methods=["POST"])
@login_required
def add_okr_key_result(objective_id):
    csrf_protect()
    db = get_db()
    objective = get_okr_or_404(objective_id)
    if not can_manage_okr(g.user, objective):
        abort(403)
    title = clean_text(request.form.get("title"), 240)
    baseline = decimal_field("baseline")
    target = decimal_field("target")
    due_date = clean_text(request.form.get("due_date"), 10)
    unit = clean_text(request.form.get("unit"), 40)
    owner_id = int_field("owner_id") or objective["owner_id"]
    if not title or baseline is None or target is None or baseline == target or not parse_report_date(due_date):
        flash("Provide a title, different numeric baseline and target, and valid due date.", "danger")
        return redirect(url_for("okr_detail", objective_id=objective_id))
    owner = db.execute("SELECT * FROM users WHERE id = ? AND deleted_at IS NULL", (owner_id,)).fetchone() if owner_id else None
    if owner_id and (not owner or (g.user["role"] == "admin" and owner["department"] != g.user["department"])):
        abort(400)
    timestamp = now()
    db.execute(
        """INSERT INTO okr_key_results
           (objective_id, title, baseline, target, current_value, unit, due_date, owner_id, status, created_at, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            objective_id, title, float(baseline), float(target), float(baseline), unit,
            due_date, owner_id, "active" if owner_id else "draft", timestamp, timestamp,
        ),
    )
    db.commit()
    flash("Key result added.", "success")
    return redirect(url_for("okr_detail", objective_id=objective_id))


@app.route("/okrs/<int:objective_id>/key-results/<int:key_result_id>/updates", methods=["POST"])
@login_required
def add_okr_update(objective_id, key_result_id):
    csrf_protect()
    db = get_db()
    objective = get_okr_or_404(objective_id)
    result = db.execute(
        "SELECT * FROM okr_key_results WHERE id = ? AND objective_id = ?", (key_result_id, objective_id)
    ).fetchone()
    if not result:
        abort(404)
    if not (can_manage_okr(g.user, objective) or objective["owner_id"] == g.user["id"] or result["owner_id"] == g.user["id"]):
        abort(403)
    if objective["status"] not in ("active", "at_risk") or result["status"] == "draft":
        flash("Approve and activate the objective and key result before recording progress.", "danger")
        return redirect(url_for("okr_detail", objective_id=objective_id))
    progress_value = decimal_field("progress_value")
    narrative = clean_text(request.form.get("narrative"), 2000)
    evidence = clean_text(request.form.get("evidence_reference"), 500)
    status = clean_text(request.form.get("status"), 20)
    if progress_value is None or not narrative or status not in OKR_KEY_RESULT_STATUSES:
        flash("Provide a numeric progress value, update note and valid status.", "danger")
        return redirect(url_for("okr_detail", objective_id=objective_id))
    timestamp = now()
    db.execute(
        """INSERT INTO okr_updates (key_result_id, user_id, progress_value, narrative, evidence_reference, created_at)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (key_result_id, g.user["id"], float(progress_value), narrative, evidence, timestamp),
    )
    db.execute(
        "UPDATE okr_key_results SET current_value = ?, status = ?, updated_at = ? WHERE id = ?",
        (float(progress_value), status, timestamp, key_result_id),
    )
    db.commit()
    flash("Progress update recorded with its evidence trail.", "success")
    return redirect(url_for("okr_detail", objective_id=objective_id))


@app.route("/okrs/<int:objective_id>/key-results/<int:key_result_id>/settings", methods=["POST"])
@login_required
def update_okr_key_result(objective_id, key_result_id):
    csrf_protect()
    db = get_db()
    objective = get_okr_or_404(objective_id)
    if not can_manage_okr(g.user, objective):
        abort(403)
    result = db.execute(
        "SELECT * FROM okr_key_results WHERE id = ? AND objective_id = ?", (key_result_id, objective_id)
    ).fetchone()
    if not result:
        abort(404)
    baseline = decimal_field("baseline")
    target = decimal_field("target")
    owner_id = int_field("owner_id") or None
    due_date = clean_text(request.form.get("due_date"), 10)
    unit = clean_text(request.form.get("unit"), 40)
    status = clean_text(request.form.get("status"), 20)
    if baseline is None or target is None or baseline == target or not parse_report_date(due_date):
        flash("Baseline and target must be different numbers and the due date must be valid.", "danger")
        return redirect(url_for("okr_detail", objective_id=objective_id))
    if status not in OKR_KEY_RESULT_STATUSES:
        abort(400)
    owner = db.execute("SELECT * FROM users WHERE id = ? AND deleted_at IS NULL", (owner_id,)).fetchone() if owner_id else None
    if owner_id and (not owner or (g.user["role"] == "admin" and owner["department"] != g.user["department"])):
        abort(400)
    if status != "draft" and not owner_id:
        flash("Assign an owner before moving a key result out of Draft.", "danger")
        return redirect(url_for("okr_detail", objective_id=objective_id))
    update_count = db.execute(
        "SELECT COUNT(*) AS c FROM okr_updates WHERE key_result_id = ?", (key_result_id,)
    ).fetchone()["c"]
    current_value = float(baseline) if not update_count else result["current_value"]
    db.execute(
        """UPDATE okr_key_results SET baseline = ?, target = ?, current_value = ?, unit = ?, due_date = ?,
           owner_id = ?, status = ?, updated_at = ? WHERE id = ?""",
        (float(baseline), float(target), current_value, unit, due_date, owner_id, status, now(), key_result_id),
    )
    db.commit()
    flash("Key result target and ownership updated.", "success")
    return redirect(url_for("okr_detail", objective_id=objective_id))


@app.route("/okrs/<int:objective_id>/status", methods=["POST"])
@login_required
def update_okr_status(objective_id):
    csrf_protect()
    objective = get_okr_or_404(objective_id)
    if not can_manage_okr(g.user, objective):
        abort(403)
    status = clean_text(request.form.get("status"), 20)
    owner_id = int_field("owner_id") or None
    if status not in OKR_STATUSES:
        abort(400)
    owner = get_db().execute("SELECT * FROM users WHERE id = ? AND deleted_at IS NULL", (owner_id,)).fetchone() if owner_id else None
    if owner_id and (not owner or (g.user["role"] == "admin" and owner["department"] != g.user["department"])):
        abort(400)
    if status == "active":
        result_summary = get_db().execute(
            """SELECT COUNT(*) AS total,
                      SUM(CASE WHEN owner_id IS NULL OR status = 'draft' THEN 1 ELSE 0 END) AS incomplete
               FROM okr_key_results WHERE objective_id = ?""",
            (objective_id,),
        ).fetchone()
        if not owner_id or not result_summary["total"] or result_summary["incomplete"]:
            flash("Assign the objective owner and every key-result owner, then move all key results out of Draft before activation.", "danger")
            return redirect(url_for("okr_detail", objective_id=objective_id))
    get_db().execute(
        "UPDATE okr_objectives SET status = ?, owner_id = ?, updated_at = ? WHERE id = ?",
        (status, owner_id, now(), objective_id),
    )
    get_db().commit()
    flash("Objective status updated.", "success")
    return redirect(url_for("okr_detail", objective_id=objective_id))


@app.route("/wingu", methods=["GET"])
@login_required
def wingu_dispatches():
    return render_template(
        "wingu_dispatches.html", dispatches=visible_wingu_dispatch_rows(),
        statuses=WINGU_DISPATCH_STATUSES, attendance_preview=None,
    )


def visible_wingu_dispatch_rows():
    rows = get_db().execute(
        """SELECT d.*, r.report_date, r.department, r.user_id, users.full_name
           FROM wingu_dispatches d JOIN reports r ON r.id = d.report_id
           JOIN users ON users.id = r.user_id ORDER BY d.queued_at DESC, d.id DESC"""
    ).fetchall()
    return [row for row in rows if can_view_report(g.user, row)]


@app.route("/wingu/attendance-template.xlsx")
@login_required
def download_wingu_attendance_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    headers = ["email", "report_date", "sign_in_time", "sign_out_time", "attendance_reference"]
    sheet.append(headers)
    sheet.append(["staff@example.com", "2026-08-31", "08:10", "17:20", "Approved attendance sheet row/reference"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="31572C")
    widths = [32, 16, 16, 16, 44]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="toolkit-wingu-attendance-template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/wingu/import-attendance", methods=["POST"])
@login_required
def import_wingu_attendance():
    csrf_protect()
    upload = request.files.get("attendance_file")
    action = clean_text(request.form.get("action"), 20)
    if action not in ("preview", "import") or not upload or not upload.filename.lower().endswith(".xlsx"):
        flash("Choose an .xlsx attendance workbook and preview or import it.", "danger")
        return redirect(url_for("wingu_dispatches"))
    safe_name = secure_filename(upload.filename) or "attendance.xlsx"
    try:
        upload_bytes = upload.stream.read()
        with zipfile.ZipFile(io.BytesIO(upload_bytes)) as archive:
            if sum(member.file_size for member in archive.infolist()) > 5 * 1024 * 1024:
                raise ValueError("Workbook expands beyond the safe processing limit")
        workbook = load_workbook(io.BytesIO(upload_bytes), read_only=True, data_only=True, keep_links=False)
        sheet = workbook.active
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    except Exception:
        flash("The attendance workbook could not be read.", "danger")
        return redirect(url_for("wingu_dispatches"))
    headers = [str(value or "").strip().lower() for value in raw_headers]
    required = ["email", "report_date", "sign_in_time", "sign_out_time", "attendance_reference"]
    if headers[:len(required)] != required:
        workbook.close()
        flash("Use the Toolkit attendance template; its five column headers must remain unchanged.", "danger")
        return redirect(url_for("wingu_dispatches"))
    preview = []
    db = get_db()
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row_number > 251:
            preview.append({"row": row_number, "status": "error", "message": "Maximum 250 data rows per import."})
            break
        if not any(value not in (None, "") for value in values[:5]):
            continue
        email = str(values[0] or "").strip().lower()[:254]
        report_date = normalise_excel_date(values[1])
        sign_in = normalise_excel_time(values[2])
        sign_out = normalise_excel_time(values[3])
        reference = clean_text(str(values[4] or ""), 300)
        item = {
            "row": row_number, "email": email, "report_date": report_date,
            "sign_in": sign_in, "sign_out": sign_out, "reference": reference,
            "status": "error", "message": "",
        }
        if not email or not report_date or not sign_in or not sign_out or sign_out <= sign_in or not reference:
            item["message"] = "Email, valid date, sign-in before sign-out, and attendance reference are required."
            preview.append(item)
            continue
        matches = db.execute(
            """SELECT reports.* FROM reports JOIN users ON users.id = reports.user_id
               WHERE lower(users.email) = ? AND reports.report_date = ? AND reports.status = 'approved'""",
            (email, report_date),
        ).fetchall()
        matches = [report for report in matches if can_view_report(g.user, report)]
        if len(matches) != 1:
            item["message"] = "Expected exactly one approved report in your access scope for this employee and date."
            preview.append(item)
            continue
        report = matches[0]
        if db.execute("SELECT id FROM wingu_dispatches WHERE report_id = ?", (report["id"],)).fetchone():
            item["message"] = "That approved report is already queued."
            preview.append(item)
            continue
        item.update({"status": "valid", "message": "Ready to queue", "report_id": report["id"]})
        preview.append(item)
    workbook.close()
    if action == "preview":
        return render_template(
            "wingu_dispatches.html", dispatches=visible_wingu_dispatch_rows(),
            statuses=WINGU_DISPATCH_STATUSES, attendance_preview=preview,
            preview_filename=safe_name,
        )
    valid_rows = [item for item in preview if item["status"] == "valid"]
    timestamp = now()
    for item in valid_rows:
        cursor = db.execute(
            """INSERT INTO wingu_dispatches
               (report_id, attendance_source, attendance_reference, sign_in_time, sign_out_time,
                status, queued_by, queued_at, updated_at)
               VALUES (?, 'excel', ?, ?, ?, 'ready', ?, ?, ?)""",
            (
                item["report_id"], f"{safe_name} · row {item['row']} · {item['reference']}",
                item["sign_in"], item["sign_out"], g.user["id"], timestamp, timestamp,
            ),
        )
        db.execute(
            "INSERT INTO wingu_dispatch_events (dispatch_id, actor_id, event, notes, created_at) VALUES (?, ?, 'queued', ?, ?)",
            (cursor.lastrowid, g.user["id"], f"Validated Excel attendance: {safe_name}, row {item['row']}", timestamp),
        )
    db.commit()
    errors = len(preview) - len(valid_rows)
    flash(f"Queued {len(valid_rows)} approved reports from Excel; {errors} rows were not imported.", "success" if valid_rows else "warning")
    return redirect(url_for("wingu_dispatches"))


@app.route("/reports/<int:report_id>/wingu", methods=["POST"])
@login_required
def queue_wingu_dispatch(report_id):
    csrf_protect()
    db = get_db()
    report = db.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
    if not report:
        abort(404)
    if not can_view_report(g.user, report) or (report["user_id"] != g.user["id"] and g.user["role"] not in ADMIN_ROLES):
        abort(403)
    if report["status"] != "approved":
        flash("Only an approved report can be queued for Wingu.", "danger")
        return redirect(url_for("view_report", report_id=report_id))
    source = clean_text(request.form.get("attendance_source"), 20)
    reference = clean_text(request.form.get("attendance_reference"), 300)
    sign_in = parse_clock_time(request.form.get("sign_in_time"))
    sign_out = parse_clock_time(request.form.get("sign_out_time"))
    if source not in ("manual", "excel") or not sign_in or not sign_out or sign_out <= sign_in:
        flash("Select manual or Excel attendance and provide a valid sign-in/sign-out range.", "danger")
        return redirect(url_for("view_report", report_id=report_id))
    if source == "excel" and not reference:
        flash("Identify the approved Excel sheet or row used as the attendance source.", "danger")
        return redirect(url_for("view_report", report_id=report_id))
    timestamp = now()
    if db.execute("SELECT id FROM wingu_dispatches WHERE report_id = ?", (report_id,)).fetchone():
        flash("This approved report is already in the Wingu queue.", "warning")
        return redirect(url_for("view_report", report_id=report_id))
    cursor = db.execute(
        """INSERT INTO wingu_dispatches
           (report_id, attendance_source, attendance_reference, sign_in_time, sign_out_time,
            status, queued_by, queued_at, updated_at)
           VALUES (?, ?, ?, ?, ?, 'ready', ?, ?, ?)""",
        (report_id, source, reference, sign_in, sign_out, g.user["id"], timestamp, timestamp),
    )
    db.execute(
        "INSERT INTO wingu_dispatch_events (dispatch_id, actor_id, event, notes, created_at) VALUES (?, ?, 'queued', ?, ?)",
        (cursor.lastrowid, g.user["id"], f"Attendance source: {source}", timestamp),
    )
    db.commit()
    flash("Approved report queued. Wingu project selection will be read from Wingu during dispatch.", "success")
    return redirect(url_for("wingu_dispatches"))


@app.route("/wingu/<int:dispatch_id>/status", methods=["POST"])
@admin_required
def update_wingu_dispatch(dispatch_id):
    csrf_protect()
    db = get_db()
    dispatch = db.execute(
        """SELECT d.*, r.user_id FROM wingu_dispatches d JOIN reports r ON r.id = d.report_id WHERE d.id = ?""",
        (dispatch_id,),
    ).fetchone()
    if not dispatch:
        abort(404)
    if not can_view_report(g.user, dispatch):
        abort(403)
    status = clean_text(request.form.get("status"), 30)
    project = clean_text(request.form.get("wingu_project"), 180)
    external_reference = clean_text(request.form.get("external_reference"), 300)
    notes = clean_text(request.form.get("notes"), 1000)
    if status not in WINGU_DISPATCH_STATUSES:
        abort(400)
    if status in ("accepted", "rejected", "needs_attention") and not notes:
        flash("Add a reconciliation note for the final Wingu result.", "danger")
        return redirect(url_for("wingu_dispatches"))
    timestamp = now()
    db.execute(
        """UPDATE wingu_dispatches SET status = ?, wingu_project = ?, external_reference = ?,
           last_error = ?, updated_at = ? WHERE id = ?""",
        (status, project, external_reference, notes if status in ("rejected", "needs_attention") else "", timestamp, dispatch_id),
    )
    db.execute(
        "INSERT INTO wingu_dispatch_events (dispatch_id, actor_id, event, notes, created_at) VALUES (?, ?, ?, ?, ?)",
        (dispatch_id, g.user["id"], status, notes, timestamp),
    )
    db.commit()
    flash("Wingu reconciliation status recorded.", "success")
    return redirect(url_for("wingu_dispatches"))


@app.route("/principal")
@principal_required
def principal_overview():
    db = get_db()
    archived_expr = archived_filter("reports.archived")
    reports = db.execute(
        f"""
        SELECT reports.*, users.full_name
        FROM reports
        JOIN users ON users.id = reports.user_id
        WHERE {archived_expr} = 0 AND users.deleted_at IS NULL
        ORDER BY reports.report_date DESC, reports.id DESC
        """
    ).fetchall()
    users = strip_hidden_users(db.execute(
        """
        SELECT id, full_name, department, position, branch, role, is_active, locked_at
        FROM users
        WHERE deleted_at IS NULL
        ORDER BY department ASC, full_name ASC
        """
    ).fetchall())
    insights = build_dashboard_insights(reports, get_visible_staff_count(db))
    overview = build_principal_overview(reports, users)
    return render_template("principal.html", insights=insights, overview=overview)




@app.route("/reports/new", methods=["GET", "POST"])
@login_required
def new_report():
    if request.method == "POST":
        csrf_protect()
        last = session.get("last_submit_at")
        if last and (datetime.now() - datetime.fromisoformat(last)).total_seconds() < RATE_LIMIT_SECONDS:
            flash("Please wait before submitting another report.", "warning")
            return redirect(url_for("dashboard"))
        form = {key: clean_text(value) for key, value in request.form.items()}
        tasks = parse_items("task", form)
        tomorrow = parse_items("tomorrow", form)
        challenges = parse_items("challenge", form)
        if not tasks:
            flash("Add at least one task or activity completed today.", "danger")
            return redirect(url_for("new_report"))

        department = clean_text(request.form.get("department")) or g.user["department"]
        dept_metrics = get_metrics_for_department(department)
        metrics = {}
        for field in dept_metrics:
            if field["type"] == "number":
                metrics[field["key"]] = int_field(field["key"])
            else:
                metrics[field["key"]] = request.form.get(field["key"], field.get("options", ["No"])[0])

        cursor = get_db().execute(
            """
            INSERT INTO reports (
                user_id, report_date, reporting_period, branch, department, position, day_summary,
                tasks_json, challenges_json, decisions, tomorrow_json, comments,
                metrics_json, status, archived, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                g.user["id"],
                request.form.get("report_date"),
                clean_text(request.form.get("reporting_period")) or default_reporting_period(request.form.get("report_date")),
                clean_text(request.form.get("branch")) or g.user["branch"],
                department,
                clean_text(request.form.get("position")) or g.user["position"],
                clean_text(request.form.get("day_summary")),
                json.dumps(tasks),
                json.dumps(challenges),
                clean_text(request.form.get("decisions")),
                json.dumps(tomorrow),
                clean_text(request.form.get("comments")),
                json.dumps(metrics),
                "submitted",
                0,
                now(),
            ),
        )
        get_db().commit()
        make_report_pdf(cursor.lastrowid)
        draft_id = request.form.get("draft_id", "")
        if draft_id:
            get_db().execute("DELETE FROM report_drafts WHERE id = ? AND user_id = ?", (draft_id, g.user["id"]))
        else:
            get_db().execute("DELETE FROM report_drafts WHERE user_id = ? AND report_date = ?", (g.user["id"], request.form.get("report_date", "")))
        get_db().commit()
        session["last_submit_at"] = now()
        flash("Report submitted and PDF generated.", "success")
        return redirect(url_for("view_report", report_id=cursor.lastrowid))

    draft_id = request.args.get("draft", "")
    today = datetime.now().strftime("%Y-%m-%d")
    return render_template(
        "report_form.html",
        departments=DEPARTMENTS,
        branches=BRANCHES,
        today=today,
        day_of_week=day_name(today),
        reporting_period=default_reporting_period(today),
        metrics_config=METRICS_CONFIG,
        load_draft_id=draft_id,
    )


@app.route("/reports/<int:report_id>")
@login_required
def view_report(report_id):
    report = get_db().execute(
        """
        SELECT reports.*, users.full_name, users.email, users.phone
        FROM reports
        JOIN users ON users.id = reports.user_id
        WHERE reports.id = ?
        """,
        (report_id,),
    ).fetchone()
    if not report:
        abort(404)
    if not can_view_report(g.user, report):
        abort(403)
    edits = get_db().execute(
        "SELECT * FROM report_edits WHERE report_id = ? ORDER BY edited_at ASC",
        (report_id,),
    ).fetchall()
    comments = get_db().execute(
        "SELECT * FROM report_comments WHERE report_id = ? ORDER BY created_at ASC",
        (report_id,),
    ).fetchall()
    return render_template(
        "report_detail.html",
        report=report,
        tasks=json.loads(report["tasks_json"]),
        challenges=json.loads(report["challenges_json"]),
        tomorrow=json.loads(report["tomorrow_json"]),
        metrics=json.loads(report["metrics_json"]),
        metrics_config=METRICS_CONFIG,
        edits=edits,
        comments=comments,
    )


@app.route("/reports/<int:report_id>/edit", methods=["GET", "POST"])
@login_required
def edit_report(report_id):
    db = get_db()
    report = db.execute(
        """
        SELECT reports.*, users.full_name, users.email, users.phone
        FROM reports
        JOIN users ON users.id = reports.user_id
        WHERE reports.id = ?
        """,
        (report_id,),
    ).fetchone()
    if not report:
        abort(404)
    if not can_view_report(g.user, report):
        abort(403)
    if report["status"] == "approved":
        flash("This report has been approved and can no longer be edited.", "danger")
        return redirect(url_for("view_report", report_id=report_id))
    if report["user_id"] != g.user["id"]:
        abort(403)

    edit_count = db.execute(
        "SELECT COUNT(*) AS cnt FROM report_edits WHERE report_id = ?", (report_id,)
    ).fetchone()["cnt"]
    if edit_count >= 3:
        flash("This report has reached the maximum of 3 edits and can no longer be edited.", "danger")
        return redirect(url_for("view_report", report_id=report_id))

    if request.method == "POST":
        csrf_protect()
        form = {key: clean_text(value) for key, value in request.form.items()}
        tasks = parse_items("task", form)
        tomorrow = parse_items("tomorrow", form)
        challenges = parse_items("challenge", form)
        if not tasks:
            flash("Add at least one task or activity completed today.", "danger")
            return redirect(url_for("edit_report", report_id=report_id))

        department = clean_text(request.form.get("department")) or report["department"]
        dept_metrics = get_metrics_for_department(department)
        metrics = {}
        for field in dept_metrics:
            if field["type"] == "number":
                metrics[field["key"]] = int_field(field["key"])
            else:
                metrics[field["key"]] = request.form.get(field["key"], field.get("options", ["No"])[0])

        db.execute(
            """
            UPDATE reports SET
                report_date = ?, reporting_period = ?, branch = ?, department = ?, position = ?,
                day_summary = ?, tasks_json = ?, challenges_json = ?, decisions = ?,
                tomorrow_json = ?, comments = ?, metrics_json = ?
            WHERE id = ?
            """,
            (
                request.form.get("report_date"),
                clean_text(request.form.get("reporting_period")) or default_reporting_period(request.form.get("report_date")),
                clean_text(request.form.get("branch")) or report["branch"],
                department,
                clean_text(request.form.get("position")) or report["position"],
                clean_text(request.form.get("day_summary")),
                json.dumps(tasks),
                json.dumps(challenges),
                clean_text(request.form.get("decisions")),
                json.dumps(tomorrow),
                clean_text(request.form.get("comments")),
                json.dumps(metrics),
                report_id,
            ),
        )
        db.execute(
            "INSERT INTO report_edits (report_id, user_id, edited_at) VALUES (?, ?, ?)",
            (report_id, g.user["id"], now()),
        )
        db.commit()
        make_report_pdf(report_id)
        flash("Report updated successfully.", "success")
        return redirect(url_for("view_report", report_id=report_id))

    return render_template(
        "report_form.html",
        report=report,
        tasks=json.loads(report["tasks_json"]),
        challenges=json.loads(report["challenges_json"]),
        tomorrow=json.loads(report["tomorrow_json"]),
        metrics=json.loads(report["metrics_json"]),
        departments=DEPARTMENTS,
        branches=BRANCHES,
        today=report["report_date"],
        day_of_week=day_name(report["report_date"]),
        reporting_period=report["reporting_period"] or default_reporting_period(report["report_date"]),
        metrics_config=METRICS_CONFIG,
    )


@app.route("/reports/<int:report_id>/pdf")
@login_required
def download_report(report_id):
    report = get_db().execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
    if not report:
        abort(404)
    if not can_view_report(g.user, report):
        abort(403)
    path = REPORT_DIR / report["pdf_filename"] if report["pdf_filename"] else make_report_pdf(report_id)
    if not path.exists():
        path = make_report_pdf(report_id)
    return send_file(path, as_attachment=True, download_name=path.name)


@app.route("/admin/users", methods=["GET", "POST"])
@superadmin_required
def admin_users():
    db = get_db()
    if request.method == "POST":
        csrf_protect()
        user_id = int(request.form.get("user_id"))
        role = request.form.get("role")
        is_active = 1 if request.form.get("is_active") == "on" else 0
        if role not in ("employee", "admin", "principal", "superadmin", "shadowadmin"):
            abort(400)
        target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        guard_admin_target(target)
        if role == "shadowadmin" and not viewer_is_owner():
            abort(403)
        if target["locked_at"] and is_active:
            is_active = 0
            flash("This account is locked. Use Unlock and record a reason before reactivating it.", "warning")
        db.execute("UPDATE users SET role = ?, is_active = ? WHERE id = ?", (role, is_active, user_id))
        db.commit()
        flash("User updated.", "success")
        return redirect(url_for("admin_users"))

    users = strip_hidden_users(db.execute("SELECT * FROM users WHERE deleted_at IS NULL ORDER BY role DESC, full_name ASC").fetchall())
    return render_template("admin_users.html", users=users)


@app.route("/admin/users/<int:user_id>/reset-password", methods=["GET", "POST"])
@superadmin_required
def admin_reset_password(user_id):
    db = get_db()
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        abort(404)
    guard_admin_target(target)
    if request.method == "POST":
        csrf_protect()
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")
        if len(password) < 8:
            flash("Use a password with at least 8 characters.", "danger")
            return render_template("admin_reset_password.html", target=target, back_url=url_for("admin_users"))
        if password != confirm:
            flash("Passwords do not match.", "danger")
            return render_template("admin_reset_password.html", target=target, back_url=url_for("admin_users"))
        db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(password), user_id))
        db.commit()
        flash(f"Password for {target['full_name']} has been reset.", "success")
        return redirect(url_for("admin_users"))
    return render_template("admin_reset_password.html", target=target, back_url=url_for("admin_users"))


@app.route("/admin/access/<int:admin_id>", methods=["GET", "POST"])
@superadmin_required
def admin_access(admin_id):
    db = get_db()
    admin = db.execute("SELECT * FROM users WHERE id = ? AND role IN ('admin', 'principal', 'superadmin')", (admin_id,)).fetchone()
    if not admin:
        abort(404)
    if request.method == "POST":
        csrf_protect()
        employee_ids = request.form.getlist("employee_ids")
        db.execute("DELETE FROM report_access WHERE admin_id = ?", (admin_id,))
        for employee_id in employee_ids:
            db.execute(
                "INSERT OR IGNORE INTO report_access (admin_id, employee_id, created_at) VALUES (?, ?, ?)",
                (admin_id, int(employee_id), now()),
            )
        db.commit()
        flash("Report access updated.", "success")
        return redirect(url_for("admin_access", admin_id=admin_id))

    employees = db.execute("SELECT * FROM users WHERE role = 'employee' ORDER BY full_name ASC").fetchall()
    assigned = {
        row["employee_id"]
        for row in db.execute("SELECT employee_id FROM report_access WHERE admin_id = ?", (admin_id,)).fetchall()
    }
    return render_template("admin_access.html", admin=admin, employees=employees, assigned=assigned)


@app.route("/admin/settings", methods=["GET", "POST"])
@superadmin_required
def admin_settings():
    db = get_db()
    if request.method == "POST":
        csrf_protect()
        registration = "1" if request.form.get("registration_open") == "on" else "0"
        db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('registration_open', ?)", (registration,))
        db.commit()
        flash("Settings saved.", "success")
        return redirect(url_for("admin_settings"))
    reg_open = db.execute("SELECT value FROM settings WHERE key = 'registration_open'").fetchone()
    return render_template("admin_settings.html", registration_open=(reg_open and reg_open["value"] == "1"))


@app.route("/admin/reminders", methods=["GET", "POST"])
@superadmin_required
def admin_reminders():
    db = get_db()
    if request.method == "POST":
        csrf_protect()
        action = clean_text(request.form.get("action"), 30)
        if action == "create":
            kind = clean_text(request.form.get("kind"), 50)
            cadence = clean_text(request.form.get("cadence"), 20)
            try:
                lead_days = int(request.form.get("lead_days", "0"))
            except (TypeError, ValueError):
                lead_days = -1
            if kind not in REMINDER_KINDS or cadence not in ("daily", "weekly") or not 0 <= lead_days <= 30:
                abort(400)
            existing = db.execute("SELECT id FROM reminder_rules WHERE kind = ?", (kind,)).fetchone()
            if existing:
                flash("A reminder rule for that workflow already exists.", "warning")
                return redirect(url_for("admin_reminders"))
            timestamp = now()
            db.execute(
                """INSERT INTO reminder_rules
                   (kind, cadence, lead_days, is_enabled, created_by, created_at, updated_at)
                   VALUES (?, ?, ?, 1, ?, ?, ?)""",
                (kind, cadence, lead_days, g.user["id"], timestamp, timestamp),
            )
            db.commit()
            flash("In-app reminder rule created.", "success")
        elif action == "toggle":
            try:
                rule_id = int(request.form.get("rule_id", ""))
            except (TypeError, ValueError):
                abort(400)
            rule = db.execute("SELECT * FROM reminder_rules WHERE id = ?", (rule_id,)).fetchone()
            if not rule:
                abort(404)
            enabled = 0 if rule["is_enabled"] else 1
            db.execute("UPDATE reminder_rules SET is_enabled = ?, updated_at = ? WHERE id = ?", (enabled, now(), rule_id))
            db.commit()
            flash("Reminder rule updated.", "success")
        else:
            abort(400)
        return redirect(url_for("admin_reminders"))
    rules = db.execute(
        """SELECT reminder_rules.*, users.full_name AS creator_name
           FROM reminder_rules JOIN users ON users.id = reminder_rules.created_by
           ORDER BY reminder_rules.id"""
    ).fetchall()
    logs = db.execute(
        """SELECT notification_delivery_logs.*, users.full_name, reminder_rules.kind
           FROM notification_delivery_logs
           JOIN users ON users.id = notification_delivery_logs.user_id
           JOIN reminder_rules ON reminder_rules.id = notification_delivery_logs.rule_id
           ORDER BY notification_delivery_logs.created_at DESC, notification_delivery_logs.id DESC
           LIMIT 100"""
    ).fetchall()
    return render_template("admin_reminders.html", rules=rules, logs=logs, reminder_kinds=REMINDER_KINDS)


@app.route("/admin/reminders/run", methods=["POST"])
@superadmin_required
def admin_run_reminders():
    csrf_protect()
    result = run_reminder_rules(force=True)
    flash(
        f"Reminder run complete: {result['delivered']} delivered, {result['skipped']} duplicate deliveries skipped.",
        "success",
    )
    return redirect(url_for("admin_reminders"))


@app.route("/admin/unlock/<int:user_id>", methods=["GET", "POST"])
@login_required
def admin_unlock(user_id):
    db = get_db()
    if g.user["role"] not in ("admin", "superadmin", "shadowadmin"):
        abort(403)
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        abort(404)
    if g.user["role"] == "admin":
        if target["role"] != "employee" or target["department"] != g.user["department"]:
            abort(404)
    else:
        guard_admin_target(target)
    if request.method == "POST":
        csrf_protect()
        reason = clean_text(request.form.get("reason", ""), 500)
        if not reason:
            flash("You must provide a reason for unlocking this account.", "danger")
            return render_template("admin_unlock.html", target=target)
        db.execute(
            "UPDATE users SET is_active = 1, locked_at = NULL, lock_reason = ? WHERE id = ?",
            (reason, user_id),
        )
        db.execute(
            """INSERT INTO account_events (user_id, actor_id, event, reason, created_at)
               VALUES (?, ?, 'unlocked', ?, ?)""",
            (user_id, g.user["id"], reason, now()),
        )
        db.commit()
        flash(f"{target['full_name']} has been unlocked. Reason recorded: {reason}", "success")
        destination = "admin_department" if g.user["role"] == "admin" else "admin_users"
        return redirect(url_for(destination))
    back_url = url_for("admin_department") if g.user["role"] == "admin" else url_for("admin_users")
    return render_template("admin_unlock.html", target=target, back_url=back_url)


@app.route("/admin/users/<int:user_id>/manage", methods=["GET", "POST"])
@superadmin_required
def admin_manage_user(user_id):
    db = get_db()
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        abort(404)
    guard_admin_target(target)
    if request.method == "POST":
        csrf_protect()
        action = request.form.get("action", "")
        data_choice = request.form.get("data_choice", "keep")
        if action == "deactivate":
            db.execute(
                "UPDATE users SET is_active = 0, data_retention = ? WHERE id = ?",
                (data_choice, user_id),
            )
            db.commit()
            flash(f"{target['full_name']} deactivated. Data: {data_choice}.", "success")
            return redirect(url_for("admin_users"))
        elif action == "delete":
            if data_choice == "delete_all":
                reports = db.execute("SELECT id, pdf_filename FROM reports WHERE user_id = ?", (user_id,)).fetchall()
                for r in reports:
                    db.execute("DELETE FROM report_edits WHERE report_id = ?", (r["id"],))
                    db.execute("DELETE FROM report_comments WHERE report_id = ?", (r["id"],))
                    if r["pdf_filename"]:
                        pdf_path = REPORT_DIR / r["pdf_filename"]
                        if pdf_path.exists():
                            pdf_path.unlink()
                db.execute("DELETE FROM reports WHERE user_id = ?", (user_id,))
            elif data_choice == "archive":
                db.execute("UPDATE reports SET archived = 1 WHERE user_id = ?", (user_id,))
            db.execute("DELETE FROM report_access WHERE admin_id = ? OR employee_id = ?", (user_id, user_id))
            db.execute("DELETE FROM report_edits WHERE user_id = ?", (user_id,))
            db.execute("DELETE FROM report_comments WHERE admin_id = ?", (user_id,))
            db.execute("DELETE FROM password_resets WHERE user_id = ?", (user_id,))
            db.execute("DELETE FROM users WHERE id = ?", (user_id,))
            db.commit()
            flash(f"{target['full_name']} deleted. Data: {data_choice}.", "success")
            return redirect(url_for("admin_users"))
        else:
            flash("Invalid action.", "danger")
            return redirect(url_for("admin_manage_user", user_id=user_id))
    report_count = db.execute("SELECT COUNT(*) AS c FROM reports WHERE user_id = ?", (user_id,)).fetchone()["c"]
    return render_template("admin_manage_user.html", target=target, report_count=report_count)


@app.route("/admin/department", methods=["GET", "POST"])
@department_admin_required
def admin_department():
    db = get_db()
    dept = g.user["department"]
    if request.method == "POST":
        csrf_protect()
        toggle_id = request.form.get("toggle_id")
        if toggle_id:
            target = db.execute("SELECT * FROM users WHERE id = ? AND department = ?", (toggle_id, dept)).fetchone()
            if target and target["role"] == "employee":
                if target["locked_at"]:
                    flash("This account is locked. Use Unlock and record a reason.", "warning")
                    return redirect(url_for("admin_department"))
                new_active = 0 if target["is_active"] else 1
                db.execute("UPDATE users SET is_active = ? WHERE id = ?", (new_active, toggle_id))
                db.commit()
                status = "activated" if new_active else "deactivated"
                flash(f"{target['full_name']} {status}.", "success")
            return redirect(url_for("admin_department"))
        full_name = clean_text(request.form.get("full_name"), 120)
        email = clean_text(request.form.get("email"), 160).lower()
        password = request.form.get("password", "")
        if len(password) < 8:
            flash("Use a password with at least 8 characters.", "danger")
            return redirect(url_for("admin_department"))
        try:
            db.execute(
                """
                INSERT INTO users (
                    full_name, email, phone, department, position, branch,
                    password_hash, role, is_active, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    full_name,
                    email,
                    clean_text(request.form.get("phone"), 40),
                    dept,
                    clean_text(request.form.get("position"), 100),
                    clean_text(request.form.get("branch")) or g.user["branch"],
                    generate_password_hash(password),
                    "employee",
                    1,
                    now(),
                ),
            )
            db.commit()
            user_id = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()["id"]
            dept_admins = db.execute(
                "SELECT id FROM users WHERE department = ? AND role IN ('admin', 'principal', 'superadmin') AND id != ?",
                (dept, user_id),
            ).fetchall()
            for da in dept_admins:
                db.execute(
                    "INSERT OR IGNORE INTO report_access (admin_id, employee_id, created_at) VALUES (?, ?, ?)",
                    (da["id"], user_id, now()),
                )
            db.commit()
            flash(f"User {full_name} added to {dept}.", "success")
        except sqlite3.IntegrityError:
            flash("That email address is already registered.", "danger")
        return redirect(url_for("admin_department"))

    users = strip_hidden_users(db.execute(
        "SELECT * FROM users WHERE department = ? ORDER BY role DESC, full_name ASC",
        (dept,),
    ).fetchall())
    return render_template("admin_department.html", dept_users=users, departments=DEPARTMENTS, branches=BRANCHES)


@app.route("/admin/department/<int:user_id>/reset-password", methods=["GET", "POST"])
@department_admin_required
def admin_department_reset_password(user_id):
    db = get_db()
    target = db.execute(
        "SELECT * FROM users WHERE id = ? AND department = ?",
        (user_id, g.user["department"]),
    ).fetchone()
    if not target:
        abort(404)
    guard_admin_target(target)
    if request.method == "POST":
        csrf_protect()
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")
        if len(password) < 8:
            flash("Use a password with at least 8 characters.", "danger")
            return render_template("admin_reset_password.html", target=target, back_url=url_for("admin_department"))
        if password != confirm:
            flash("Passwords do not match.", "danger")
            return render_template("admin_reset_password.html", target=target, back_url=url_for("admin_department"))
        db.execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(password), user_id))
        db.commit()
        flash(f"Password for {target['full_name']} has been reset.", "success")
        return redirect(url_for("admin_department"))
    return render_template("admin_reset_password.html", target=target, back_url=url_for("admin_department"))


@app.route("/reports/<int:report_id>/status", methods=["POST"])
@login_required
def update_status(report_id):
    csrf_protect()
    if g.user["role"] not in ADMIN_ROLES:
        abort(403)
    report = get_db().execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
    if not report:
        abort(404)
    if not can_view_report(g.user, report):
        abort(403)
    new_status = request.form.get("status", "")
    if new_status not in ("submitted", "reviewed", "approved"):
        abort(400)
    get_db().execute("UPDATE reports SET status = ? WHERE id = ?", (new_status, report_id))
    get_db().commit()
    if new_status in ("reviewed", "approved") and report["user_id"] != g.user["id"]:
        user = get_db().execute("SELECT full_name FROM users WHERE id = ?", (g.user["id"],)).fetchone()
        actor = user["full_name"] if user else "An administrator"
        create_notification(
            user_id=report["user_id"],
            kind="status",
            title=f"Report {new_status}",
            body=f"{actor} marked your report from {report['report_date']} as {new_status}.",
            link=url_for("view_report", report_id=report_id),
            actor_id=g.user["id"],
        )
    flash(f"Report marked as {new_status}.", "success")
    return redirect(url_for("view_report", report_id=report_id))


@app.route("/reports/export.csv")
@login_required
def export_csv():
    db = get_db()
    rows = apply_report_filters(visible_report_rows(db, 0), report_filter_values())

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Employee", "Department", "Position", "Branch", "Summary", "Status", "Submitted At"])
    for r in rows:
        writer.writerow([
            csv_safe(r["report_date"]),
            csv_safe(r["full_name"]),
            csv_safe(r["department"]),
            csv_safe(r["position"]),
            csv_safe(r["branch"]),
            csv_safe(r["day_summary"]),
            csv_safe(r["status"]),
            csv_safe(r["created_at"]),
        ])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=toolkit_reports_export.csv"},
    )


@app.route("/reports/export.xlsx")
@login_required
def export_xlsx():
    rows = apply_report_filters(visible_report_rows(get_db(), 0), report_filter_values())
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Toolkit Reports"
    headers = ["Date", "Employee", "Department", "Position", "Branch", "Summary", "Status", "Submitted At"]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1B4D2E")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for report in rows:
        sheet.append([
            csv_safe(report["report_date"]),
            csv_safe(report["full_name"]),
            csv_safe(report["department"]),
            csv_safe(report["position"]),
            csv_safe(report["branch"]),
            csv_safe(report["day_summary"]),
            csv_safe(report["status"]),
            csv_safe(report["created_at"]),
        ])
    widths = {"A": 14, "B": 24, "C": 22, "D": 24, "E": 30, "F": 58, "G": 14, "H": 22}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="toolkit_reports_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/reports/<int:report_id>/archive", methods=["POST"])
@login_required
def archive_report(report_id):
    csrf_protect()
    report = get_db().execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
    if not report:
        abort(404)
    if not can_view_report(g.user, report):
        abort(403)
    archived = 1 if request.form.get("archive") == "1" else 0
    get_db().execute("UPDATE reports SET archived = ? WHERE id = ?", (archived, report_id))
    get_db().commit()
    flash("Report archived." if archived else "Report unarchived.", "success")
    return redirect(url_for("view_report", report_id=report_id))


@app.route("/reports/<int:report_id>/delete", methods=["POST"])
@superadmin_required
def delete_report(report_id):
    csrf_protect()
    db = get_db()
    report = db.execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
    if not report:
        abort(404)
    pdf = report["pdf_filename"]
    db.execute("DELETE FROM report_edits WHERE report_id = ?", (report_id,))
    db.execute("DELETE FROM report_comments WHERE report_id = ?", (report_id,))
    db.execute("DELETE FROM reports WHERE id = ?", (report_id,))
    db.commit()
    if pdf:
        pdf_path = REPORT_DIR / pdf
        if pdf_path.exists():
            pdf_path.unlink()
    flash("Report permanently deleted.", "success")
    return redirect(url_for("dashboard"))


@app.route("/reports/<int:report_id>/comment", methods=["POST"])
@login_required
def add_comment(report_id):
    if g.user["role"] not in ADMIN_ROLES:
        abort(403)
    csrf_protect()
    report = get_db().execute("SELECT * FROM reports WHERE id = ?", (report_id,)).fetchone()
    if not report:
        abort(404)
    if not can_view_report(g.user, report):
        abort(403)
    comment = clean_text(request.form.get("comment"), 2000)
    if not comment:
        flash("Comment cannot be empty.", "danger")
        return redirect(url_for("view_report", report_id=report_id))
    get_db().execute(
        "INSERT INTO report_comments (report_id, admin_id, admin_name, comment, created_at) VALUES (?, ?, ?, ?, ?)",
        (report_id, g.user["id"], g.user["full_name"], comment, now()),
    )
    get_db().commit()
    if report["user_id"] != g.user["id"]:
        create_notification(
            user_id=report["user_id"],
            kind="comment",
            title="New comment on your report",
            body=f"{g.user['full_name']} commented on your report from {report['report_date']}: {comment[:120]}{'...' if len(comment) > 120 else ''}",
            link=url_for("view_report", report_id=report_id),
            actor_id=g.user["id"],
        )
    flash("Comment added.", "success")
    return redirect(url_for("view_report", report_id=report_id))


def cleanup_expired_resets():
    try:
        db = get_db()
        db.execute("DELETE FROM password_resets WHERE expires_at < ?", (now(),))
        db.execute("DELETE FROM notifications WHERE is_read = 1 AND created_at < ?", (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S"))
        db.commit()
    except Exception:
        pass


@app.cli.command("reminders-run")
def reminders_run_command():
    """Deliver due in-app reminder rules; safe for a scheduled cron invocation."""
    result = run_reminder_rules(force=False)
    print(json.dumps(result, sort_keys=True))


@app.route("/health")
def health_check():
    db_ok = False
    try:
        get_db().execute("SELECT 1")
        db_ok = True
    except Exception:
        pass
    status = 200 if db_ok else 503
    return {"status": "healthy" if db_ok else "unhealthy", "database": "connected" if db_ok else "disconnected"}, status


@app.errorhandler(403)
def forbidden(_error):
    return render_template("error.html", title="Access restricted", message="You do not have permission to view this page or report."), 403


@app.errorhandler(400)
def bad_request(_error):
    return render_template("error.html", title="Bad request", message="The request could not be processed. Try reloading and submitting again."), 400


@app.errorhandler(404)
def not_found(_error):
    return render_template("error.html", title="Not found", message="The requested page or report was not found."), 404


@app.errorhandler(500)
def server_error(_error):
    log.exception("Unhandled server error")
    try:
        return render_template("error.html", title="Server error", message="Something went wrong. Please try again or contact the administrator."), 500
    except Exception:
        return "Internal Server Error", 500, {"Content-Type": "text/plain"}


init_db()
cleanup_expired_resets()
validate_production_config()
if app.config["SECRET_KEY"] == "local-dev-change-me":
    import warnings
    warnings.warn("Using insecure default SECRET_KEY. Set the SECRET_KEY environment variable for production.")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5055"))
    debug = os.environ.get("DEBUG", "").lower() in ("1", "true", "yes")
    app.run(host=os.environ.get("HOST", "0.0.0.0"), port=port, debug=debug)
