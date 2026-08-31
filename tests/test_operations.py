from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

import app as app_module


def login_session(client, user_id=1):
    with client.session_transaction() as sess:
        sess["user_id"] = user_id
        sess["_created_at"] = "2026-08-28 00:00:00"
        sess["_csrf_token"] = "test-token"
        sess.permanent = True


def add_employee(email="officer@example.test", department="Admissions", active=1, locked_at=None):
    db = app_module.get_db()
    cursor = db.execute(
        """INSERT INTO users
           (full_name, email, department, position, branch, password_hash, role,
            is_active, created_at, last_login, locked_at, lock_reason)
           VALUES (?, ?, ?, ?, ?, ?, 'employee', ?, ?, ?, ?, ?)""",
        (
            "Admissions Officer",
            email,
            department,
            "Admissions Officer",
            "Toolkit Africa Main Office",
            generate_password_hash("password123"),
            active,
            app_module.now(),
            (datetime.now() - timedelta(days=15)).strftime("%Y-%m-%d %H:%M:%S"),
            locked_at,
            "Auto-locked: inactive for 14+ days" if locked_at else "",
        ),
    )
    db.commit()
    return cursor.lastrowid


def add_user(email, role, department="Management"):
    user_id = add_employee(email=email, department=department)
    app_module.get_db().execute("UPDATE users SET role = ? WHERE id = ?", (role, user_id))
    app_module.get_db().commit()
    return user_id


class TestIntakeTargets:
    def test_operational_module_pages_render_for_admin(self, client):
        login_session(client)
        expected = {
            "/admissions": b"Admission records",
            "/intake-targets": b"Monthly intake targets",
            "/meetings": b"Meetings and actions",
            "/modules": b"Operational workflow",
        }
        for path, marker in expected.items():
            response = client.get(path)
            assert response.status_code == 200
            assert marker in response.data

    def test_target_uses_verified_admissions_for_selected_month(self, client):
        login_session(client)
        response = client.post(
            "/intake-targets",
            data={
                "_csrf_token": "test-token",
                "officer_id": "1",
                "target_month": "2026-08",
                "target_count": "2",
            },
            follow_redirects=False,
        )
        assert response.status_code == 302
        with app_module.app.app_context():
            db = app_module.get_db()
            db.execute(
                """INSERT INTO admissions
                   (applicant_name, course, admission_date, fee_status, status, created_by, created_at, updated_at)
                   VALUES (?, ?, ?, 'paid', 'verified', ?, ?, ?)""",
                ("Fee-paid Learner", "Welding", "2026-08-15", 1, app_module.now(), app_module.now()),
            )
            db.execute(
                """INSERT INTO admissions
                   (applicant_name, course, admission_date, fee_status, status, created_by, created_at, updated_at)
                   VALUES (?, ?, ?, 'unpaid', 'verified', ?, ?, ?)""",
                ("Unpaid Learner", "Welding", "2026-08-16", 1, app_module.now(), app_module.now()),
            )
            db.commit()

        page = client.get("/intake-targets?month=2026-08")
        assert page.status_code == 200
        assert b"Verified actual" in page.data
        assert b"50%" in page.data

    def test_employee_cannot_set_intake_target(self, client):
        with app_module.app.app_context():
            officer_id = add_employee()
        login_session(client, officer_id)
        response = client.post(
            "/intake-targets",
            data={
                "_csrf_token": "test-token",
                "officer_id": str(officer_id),
                "target_month": "2026-08",
                "target_count": "10",
            },
        )
        assert response.status_code == 403

    def test_fee_paid_state_requires_reference_and_is_audited(self, client):
        login_session(client)
        with app_module.app.app_context():
            db = app_module.get_db()
            admission_id = db.execute(
                """INSERT INTO admissions
                   (applicant_name, course, status, fee_status, created_by, created_at, updated_at)
                   VALUES ('Paid Learner', 'Plumbing', 'verified', 'unpaid', 1, ?, ?)""",
                (app_module.now(), app_module.now()),
            ).lastrowid
            db.commit()
        missing_reference = client.post(
            f"/admissions/{admission_id}/payment",
            data={"_csrf_token": "test-token", "action": "mark_paid", "payment_reference": ""},
            follow_redirects=True,
        )
        assert b"payment reference" in missing_reference.data
        response = client.post(
            f"/admissions/{admission_id}/payment",
            data={
                "_csrf_token": "test-token",
                "action": "mark_paid",
                "payment_reference": "RCPT-2026-001",
                "note": "Finance receipt checked.",
            },
            follow_redirects=False,
        )
        assert response.status_code == 302
        with app_module.app.app_context():
            admission = app_module.get_db().execute("SELECT * FROM admissions WHERE id = ?", (admission_id,)).fetchone()
            history = app_module.get_db().execute(
                "SELECT * FROM admission_verification_history WHERE admission_id = ? ORDER BY id DESC",
                (admission_id,),
            ).fetchone()
        assert admission["fee_status"] == "paid"
        assert admission["fee_payment_reference"] == "RCPT-2026-001"
        assert history["status"] == "fee_paid"


class TestMeetings:
    def test_meeting_action_can_be_assigned_and_completed(self, client):
        login_session(client)
        response = client.post(
            "/meetings",
            data={
                "_csrf_token": "test-token",
                "title": "Admissions follow-up",
                "meeting_date": "2026-08-28",
                "department": "Admissions",
                "attendees": "Admissions team",
                "summary": "Agreed to follow up verified applicants.",
            },
            follow_redirects=False,
        )
        assert response.status_code == 302
        with app_module.app.app_context():
            meeting = app_module.get_db().execute("SELECT * FROM meetings").fetchone()

        response = client.post(
            f"/meetings/{meeting['id']}/actions",
            data={
                "_csrf_token": "test-token",
                "description": "Call verified applicants",
                "owner_id": "1",
                "due_date": "2026-08-29",
            },
            follow_redirects=False,
        )
        assert response.status_code == 302
        with app_module.app.app_context():
            action = app_module.get_db().execute("SELECT * FROM meeting_actions").fetchone()

        response = client.post(
            f"/meetings/{meeting['id']}/actions/{action['id']}/status",
            data={
                "_csrf_token": "test-token",
                "status": "completed",
                "completion_notes": "All calls completed.",
            },
            follow_redirects=False,
        )
        assert response.status_code == 302
        with app_module.app.app_context():
            updated = app_module.get_db().execute("SELECT * FROM meeting_actions WHERE id = ?", (action["id"],)).fetchone()
        assert updated["status"] == "completed"
        assert updated["completed_at"]

    def test_blocked_action_requires_explanation(self, client):
        login_session(client)
        with app_module.app.app_context():
            db = app_module.get_db()
            meeting_id = db.execute(
                """INSERT INTO meetings
                   (title, meeting_date, department, summary, created_by, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                ("Review", "2026-08-28", "Management", "Review actions", 1, app_module.now(), app_module.now()),
            ).lastrowid
            action_id = db.execute(
                """INSERT INTO meeting_actions
                   (meeting_id, description, owner_id, status, created_by, created_at, updated_at)
                   VALUES (?, ?, ?, 'open', ?, ?, ?)""",
                (meeting_id, "Prepare evidence", 1, 1, app_module.now(), app_module.now()),
            ).lastrowid
            db.commit()
        response = client.post(
            f"/meetings/{meeting_id}/actions/{action_id}/status",
            data={"_csrf_token": "test-token", "status": "blocked", "completion_notes": ""},
            follow_redirects=True,
        )
        assert b"Explain what is blocking" in response.data


class TestAccountUnlock:
    def test_principal_cannot_manage_or_unlock_accounts(self, client):
        with app_module.app.app_context():
            db = app_module.get_db()
            db.execute("UPDATE users SET role = 'principal' WHERE id = 1")
            employee_id = add_employee(locked_at=app_module.now(), active=0)
        login_session(client)
        assert client.get("/admin/department").status_code == 403
        assert client.get(f"/admin/unlock/{employee_id}").status_code == 403

    def test_department_admin_can_unlock_employee_without_role_change(self, client):
        with app_module.app.app_context():
            db = app_module.get_db()
            db.execute("UPDATE users SET role = 'admin', department = 'Admissions' WHERE id = 1")
            employee_id = add_employee(locked_at=app_module.now(), active=0)
        login_session(client)
        response = client.post(
            f"/admin/unlock/{employee_id}",
            data={"_csrf_token": "test-token", "reason": "Employee identity confirmed by department lead."},
            follow_redirects=False,
        )
        assert response.status_code == 302
        assert "/admin/department" in response.headers["Location"]
        with app_module.app.app_context():
            user = app_module.get_db().execute("SELECT * FROM users WHERE id = ?", (employee_id,)).fetchone()
            event = app_module.get_db().execute(
                "SELECT * FROM account_events WHERE user_id = ? ORDER BY id DESC",
                (employee_id,),
            ).fetchone()
        assert user["role"] == "employee"
        assert user["is_active"] == 1
        assert user["locked_at"] is None
        assert event["event"] == "unlocked"
        assert event["actor_id"] == 1

    def test_promoting_locked_user_does_not_bypass_lock(self, client):
        with app_module.app.app_context():
            employee_id = add_employee(locked_at=app_module.now(), active=0)
        login_session(client)
        response = client.post(
            "/admin/users",
            data={
                "_csrf_token": "test-token",
                "user_id": str(employee_id),
                "role": "admin",
                "is_active": "on",
            },
            follow_redirects=False,
        )
        assert response.status_code == 302
        with app_module.app.app_context():
            user = app_module.get_db().execute("SELECT * FROM users WHERE id = ?", (employee_id,)).fetchone()
            app_module.init_db()
            refreshed = app_module.get_db().execute("SELECT * FROM users WHERE id = ?", (employee_id,)).fetchone()
        assert user["is_active"] == 0
        assert user["locked_at"] is not None
        assert refreshed["locked_at"] is not None

        with client.session_transaction() as sess:
            sess.clear()
            sess["_csrf_token"] = "test-token"
        login = client.post(
            "/login",
            data={
                "_csrf_token": "test-token",
                "email": "officer@example.test",
                "password": "password123",
            },
            follow_redirects=True,
        )
        assert b"locked due to inactivity" in login.data

    def test_fourteen_day_inactivity_creates_auditable_lock(self, client):
        with app_module.app.app_context():
            employee_id = add_employee(active=1)
        with client.session_transaction() as sess:
            sess["_csrf_token"] = "test-token"
        response = client.post(
            "/login",
            data={
                "_csrf_token": "test-token",
                "email": "officer@example.test",
                "password": "password123",
            },
            follow_redirects=True,
        )
        assert b"locked due to 14 days of inactivity" in response.data
        with app_module.app.app_context():
            user = app_module.get_db().execute("SELECT * FROM users WHERE id = ?", (employee_id,)).fetchone()
            event = app_module.get_db().execute(
                "SELECT * FROM account_events WHERE user_id = ? ORDER BY id DESC",
                (employee_id,),
            ).fetchone()
        assert user["is_active"] == 0
        assert user["locked_at"] is not None
        assert event["event"] == "auto_locked"


class TestReportFiltersAndExcel:
    def test_dashboard_and_exports_share_filters(self, client):
        login_session(client)
        with app_module.app.app_context():
            db = app_module.get_db()
            for report_date, department, status, summary in (
                ("2026-08-20", "ICT", "approved", "Keep this approved ICT report"),
                ("2026-08-21", "Marketing", "submitted", "Exclude this marketing report"),
            ):
                db.execute(
                    """INSERT INTO reports
                       (user_id, report_date, branch, department, position, day_summary,
                        tasks_json, challenges_json, tomorrow_json, metrics_json, status, archived, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, '[]', '[]', '[]', '{}', ?, 0, ?)""",
                    (1, report_date, "Toolkit Africa Main Office", department, "Officer", summary, status, app_module.now()),
                )
            db.commit()

        query = "department=ICT&status=approved&date_from=2026-08-01&date_to=2026-08-31"
        dashboard = client.get(f"/dashboard?{query}")
        assert dashboard.status_code == 200
        assert b"Keep this approved ICT report" in dashboard.data
        assert b"Exclude this marketing report" not in dashboard.data

        csv_response = client.get(f"/reports/export.csv?{query}")
        assert csv_response.status_code == 200
        assert b"Keep this approved ICT report" in csv_response.data
        assert b"Exclude this marketing report" not in csv_response.data

        xlsx_response = client.get(f"/reports/export.xlsx?{query}")
        assert xlsx_response.status_code == 200
        workbook = load_workbook(BytesIO(xlsx_response.data), read_only=True)
        values = list(workbook.active.values)
        assert values[0][0] == "Date"
        assert len(values) == 2
        assert values[1][5] == "Keep this approved ICT report"

    def test_excel_export_neutralises_formula_cells(self, client):
        login_session(client)
        with app_module.app.app_context():
            db = app_module.get_db()
            db.execute(
                """INSERT INTO reports
                   (user_id, report_date, branch, department, position, day_summary,
                    tasks_json, challenges_json, tomorrow_json, metrics_json, status, archived, created_at)
                   VALUES (1, '2026-08-28', 'Main', 'ICT', 'Officer', '=2+2',
                           '[]', '[]', '[]', '{}', 'submitted', 0, ?)""",
                (app_module.now(),),
            )
            db.commit()
        response = client.get("/reports/export.xlsx")
        workbook = load_workbook(BytesIO(response.data), read_only=True, data_only=False)
        assert list(workbook.active.values)[1][5] == "'=2+2"


class TestIncentives:
    def test_proposal_approval_and_payment_are_separate(self, client):
        with app_module.app.app_context():
            admin_id = add_user("dept-admin@example.test", "admin", "Admissions")
            principal_id = add_user("principal@example.test", "principal", "Management")
        login_session(client, admin_id)
        proposal = client.post(
            "/incentives",
            data={
                "_csrf_token": "test-token",
                "employee_id": str(admin_id),
                "period_month": "2026-08",
                "units": "3",
                "rate": "125.50",
                "description": "Three verified enrolments",
                "notes": "Evidence reviewed by department.",
            },
            follow_redirects=False,
        )
        assert proposal.status_code == 302
        with app_module.app.app_context():
            incentive = app_module.get_db().execute("SELECT * FROM incentives").fetchone()
        assert incentive["amount_cents"] == 37650
        assert incentive["status"] == "proposed"

        login_session(client, principal_id)
        approval = client.post(
            f"/incentives/{incentive['id']}/status",
            data={"_csrf_token": "test-token", "action": "approve", "notes": "Approved against verified evidence."},
            follow_redirects=False,
        )
        assert approval.status_code == 302

        login_session(client, 1)
        payment = client.post(
            f"/incentives/{incentive['id']}/status",
            data={
                "_csrf_token": "test-token",
                "action": "mark_paid",
                "payment_reference": "PAY-2026-0001",
                "notes": "Finance payment recorded.",
            },
            follow_redirects=False,
        )
        assert payment.status_code == 302
        with app_module.app.app_context():
            updated = app_module.get_db().execute("SELECT * FROM incentives WHERE id = ?", (incentive["id"],)).fetchone()
            events = app_module.get_db().execute(
                "SELECT event FROM incentive_events WHERE incentive_id = ? ORDER BY id",
                (incentive["id"],),
            ).fetchall()
        assert updated["status"] == "paid"
        assert updated["payment_reference"] == "PAY-2026-0001"
        assert [row["event"] for row in events] == ["proposed", "approved", "paid"]

    def test_proposal_creator_cannot_self_approve(self, client):
        login_session(client)
        with app_module.app.app_context():
            db = app_module.get_db()
            incentive_id = db.execute(
                """INSERT INTO incentives
                   (employee_id, period_month, description, units, rate_cents, amount_cents,
                    status, created_by, created_at, updated_at)
                   VALUES (1, '2026-08', 'Test', 1, 10000, 10000, 'proposed', 1, ?, ?)""",
                (app_module.now(), app_module.now()),
            ).lastrowid
            db.commit()
        response = client.post(
            f"/incentives/{incentive_id}/status",
            data={"_csrf_token": "test-token", "action": "approve", "notes": "Self approval"},
            follow_redirects=True,
        )
        assert b"different authorised executive" in response.data


class TestReminders:
    def test_in_app_reminder_is_logged_and_idempotent(self, client):
        login_session(client)
        with app_module.app.app_context():
            db = app_module.get_db()
            db.execute(
                """INSERT INTO admissions
                   (applicant_name, course, status, created_by, created_at, updated_at)
                   VALUES ('Pending Learner', 'Welding', 'pending', 1, ?, ?)""",
                (app_module.now(), app_module.now()),
            )
            rule_id = db.execute(
                """INSERT INTO reminder_rules
                   (kind, cadence, lead_days, is_enabled, created_by, created_at, updated_at)
                   VALUES ('admission_pending', 'daily', 0, 1, 1, ?, ?)""",
                (app_module.now(), app_module.now()),
            ).lastrowid
            db.commit()
            first = app_module.run_reminder_rules(force=True)
            second = app_module.run_reminder_rules(force=True)
            logs = db.execute("SELECT * FROM notification_delivery_logs WHERE rule_id = ?", (rule_id,)).fetchall()
            notifications = db.execute("SELECT * FROM notifications WHERE kind = 'reminder'").fetchall()
        assert first["delivered"] == 1
        assert second["delivered"] == 0
        assert second["skipped"] == 1
        assert len(logs) == 1
        assert len(notifications) == 1

    def test_only_superadmin_can_configure_reminders(self, client):
        with app_module.app.app_context():
            employee_id = add_employee()
        login_session(client, employee_id)
        assert client.get("/admin/reminders").status_code == 403
        login_session(client, 1)
        page = client.get("/admin/reminders")
        assert page.status_code == 200
        assert b"In-app reminders" in page.data
